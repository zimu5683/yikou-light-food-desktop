from playwright.sync_api import sync_playwright, TimeoutError
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import os
import re
import datetime 
import threading
import sys

# ===================== 核心配置 =====================
TARGET_URL = "https://m.icall.me/admin/#/login"
PHONE_NUMBER = ""
LOGIN_PASSWORD = ""
EXCEL_PATH = ""
ADDRESS_SHEET_MAP = {"联建": "衣锦", "衣锦": "衣锦", "医学院": "医学院", "东湖": "东湖"}
EXTENDED_ADDRESS_MAP = {"农林": "东湖"}
WEEKDAY_SHEET_COL_MAP = {
    "中餐": {"单号": "A", "姓名": "B", "地址": "C", "电话": "D", "经济/豪华": "E", "总餐次": "F"},
    "晚餐": {"单号": "G", "姓名": "H", "地址": "I", "电话": "J", "经济/豪华": "K", "总餐次": "L"}
}
ADDRESS_MEAL_COL_MAP = {
    "单号": "A", "姓名": "B", "地址": "C", "电话": "D",
    "周一": "E", "周二": "F", "周三": "G", "周四": "H",
    "周五": "I", "周六": "J", "周日": "K", "类型": "L",
    "经济/豪华": "M", "总餐次": "N"
}
WEEKDAY_TO_SHEET = {0: "周一", 1: "周二", 2: "周三", 3: "周四", 4: "周五", 5: "周六", 6: "周日"}
WEEKDAY_INDEX_TO_NAME = {0: "周一", 1: "周二", 2: "周三", 3: "周四", 4: "周五", 5: "周六", 6: "周日"}

# 性能配置
SHORT_WAIT = 500       # 改大：详情返回后等待 500ms，确保表格渲染完成
MEDIUM_WAIT = 300      # 详情加载等待
MAX_PAGE_SEARCH = 20
ELEMENT_TIMEOUT = 8000
NETWORK_IDLE_TIMEOUT = 5000
PAGE_JUMP_THRESHOLD = 6
ORDER_SEARCH_TIMEOUT = 1000
RETRY_WAIT = 300       # 搜索失败重试前等待

# ===================== 预编译正则（保持不变） =====================
REG_RECEIVER_BRACKET = re.compile(r'^\s*(.+?)\s*[（(](\d+)[）)]\s*$')
REG_NUMBERS = re.compile(r'\d+')
REG_DAXI = re.compile(r'大西.*?([a-zA-Z]+\d+|\d+[a-zA-Z]+)', re.IGNORECASE)
REG_XIAOXI = re.compile(r'小西.*?([a-zA-Z]+\d+|\d+[a-zA-Z]+)', re.IGNORECASE)
REG_YIXUEYUAN_ADDR = re.compile(r'([a-zA-Z\u4e00-\u9fa5\d]+号楼|[^\x00-\xff]+)$')
REG_LAST_SEGMENT = re.compile(r'([a-zA-Z\u4e00-\u9fa5\d]+)$')
REG_MEAL_COUNT = re.compile(r'x\s*(\d+)', re.IGNORECASE)
REG_MEAL_SPLIT = re.compile(r'（午餐）|（晚餐）')

# ===================== 工具函数（全部保留） =====================
def save_excel_with_retry(wb, excel_path):
    while True:
        try:
            wb.save(excel_path)
            print(f"Excel保存成功：{excel_path}")
            return True
        except PermissionError:
            input("\nExcel文件被占用！请关闭Excel后按回车重试...")
            print("重新保存Excel...")
        except Exception as e:
            print(f"\nExcel保存失败：{str(e)}")
            input("按回车退出重试...")
            return False

def clear_sheet_from_row3(sheet):
    if sheet.max_row < 3:
        print(f"表格[{sheet.title}]无第3行及以后内容，无需清空")
        return
    for row in range(3, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = None
    print(f"已清空表格[{sheet.title}]第3行及以后的内容")

def split_text_and_number(text):
    if not text:
        return "", ""
    number_part = ''.join(REG_NUMBERS.findall(text))
    text_part = REG_NUMBERS.sub('', text).strip()
    return text_part, number_part

def parse_receiver_info(receiver_text):
    if not receiver_text:
        return "", ""
    bracket_match = REG_RECEIVER_BRACKET.match(receiver_text)
    if bracket_match:
        name = bracket_match.group(1).strip()
        phone = bracket_match.group(2).strip()
        return name, phone
    name, phone = split_text_and_number(receiver_text)
    return name, phone

def cache_merged_ranges(sheet):
    return set(cell for merged_range in sheet.merged_cells.ranges for cell in merged_range)

def is_merged_cell(merged_ranges, cell_coord):
    return cell_coord in merged_ranges

def get_first_empty_row(sheet, merged_ranges, start_col="A"):
    last_row = sheet.max_row
    for row in range(last_row, 0, -1):
        cell_coord = f"{start_col}{row}"
        if cell_coord in merged_ranges:
            continue
        if sheet[cell_coord].value not in [None, "", " "]:
            return row + 1
    return 1

def get_weekday_fill_value():
    current_weekday = datetime.datetime.now().weekday()
    next_weekday = 0 if current_weekday == 6 else current_weekday + 1
    target_weekday = WEEKDAY_INDEX_TO_NAME[next_weekday]
    fill_dict = {name: "" for name in WEEKDAY_INDEX_TO_NAME.values()}
    fill_dict[target_weekday] = 1
    return fill_dict

def extract_text_after_label(page, label):
    try:
        label_elem = page.locator(f'text={label}').first
        label_elem.wait_for(timeout=ELEMENT_TIMEOUT)
        full_text = label_elem.locator('..').inner_text().strip()
        content = full_text.split(label)[-1].strip().lstrip('：:').strip()
        return content
    except:
        try:
            content = page.locator(f'//*[text()="{label}"]/following-sibling::*').first.inner_text().strip()
            return content.lstrip('：:').strip()
        except:
            return ""

def extract_meal_info(page, meal_type):
    meal_info_list = []
    try:
        rows = page.eval_on_selector_all('.table_box tbody tr', '''rows => rows.map(r => {
            const prod = (r.querySelector('td:nth-child(1)') && r.querySelector('td:nth-child(1)').innerText) || '';
            const qty = (r.querySelector('td:nth-child(3)') && r.querySelector('td:nth-child(3)').innerText) || '';
            return {product: prod.trim(), qty: qty.trim()};
        })''')
        for row in rows:
            product_full_text = row.get('product') if isinstance(row, dict) else row['product']
            quantity_text = row.get('qty') if isinstance(row, dict) else row['qty']
            meal_segments = REG_MEAL_SPLIT.split(product_full_text)
            split_chars = REG_MEAL_SPLIT.findall(product_full_text)
            valid_segments = []
            for i in range(max(0, len(meal_segments)-1)):
                segment_content = meal_segments[i].strip()
                split_char = split_chars[i] if i < len(split_chars) else ''
                segment_meal = "午餐" if split_char == "（午餐）" else "晚餐"
                if segment_content:
                    valid_segments.append((segment_content, segment_meal))
            for seg_content, seg_meal in valid_segments:
                if seg_meal != meal_type:
                    continue
                full_meal_name = f"{seg_content}（{seg_meal}）"
                meal_count = int(REG_MEAL_COUNT.search(quantity_text).group(1)) if REG_MEAL_COUNT.search(quantity_text) else 1
                if "六餐" in full_meal_name:
                    total_meals = 6
                elif "单点" in full_meal_name:
                    total_meals = 1
                else:
                    total_meals = None
                meal_grade = "经济" if "经济" in full_meal_name else "豪华" if "豪华" in full_meal_name else None
                meal_info = {
                    "总餐次": total_meals,
                    "经济/豪华": meal_grade,
                    "count": meal_count
                }
                meal_info_list.append(meal_info)
    except Exception as e:
        print(f"提取{meal_type}失败：{str(e)}")
        import traceback
        traceback.print_exc()
    return meal_info_list

def extract_product_note_text(page):
    try:
        product_texts = page.eval_on_selector_all('.table_box tbody tr', '''rows => rows.map(r => {
            const prod = (r.querySelector('td:nth-child(1)') && r.querySelector('td:nth-child(1)').innerText) || '';
            return prod.trim();
        })''')
        note_texts = []
        for product_text in product_texts:
            lines = [line.strip() for line in product_text.splitlines() if line.strip()]
            if len(lines) > 1:
                note_texts.extend(lines[1:])
        return " ".join(note_texts)
    except Exception as e:
        print(f"提取商品备注失败：{str(e)}")
        return ""

def get_yijin_address_from_product_note(product_note):
    return "外卖柜" if "联建门口外卖柜" in product_note else "校门口"

def jump_to_page(page, target_page):
    if target_page == 1:
        return True
    if target_page > MAX_PAGE_SEARCH:
        print(f"超过最大查找页码{MAX_PAGE_SEARCH}，停止查找")
        return False
    pages_to_click = []
    if target_page <= PAGE_JUMP_THRESHOLD:
        pages_to_click.append(target_page)
    else:
        step = PAGE_JUMP_THRESHOLD
        while step < target_page:
            pages_to_click.append(step)
            step += 2
        pages_to_click.append(target_page)
    if len(pages_to_click) > 1:
        print(f"目标页{target_page}大于阈值，将依次跳转：{pages_to_click}")
    for page_num in pages_to_click:
        try:
            page_btn = page.locator(
                f'//div[contains(@class,"el-pagination")]//li[contains(@class,"number") and text()="{page_num}"]'
            ).first
            page_btn.wait_for(timeout=ELEMENT_TIMEOUT)
            page_btn.click()
            page.wait_for_load_state("domcontentloaded", timeout=NETWORK_IDLE_TIMEOUT)
            page.wait_for_timeout(SHORT_WAIT)
        except Exception as e:
            print(f"无法跳转到第{page_num}页：{str(e)}")
            return False
    return True

def get_donghu_address_segment(delivery_address):
    daxi_match = REG_DAXI.search(delivery_address)
    if daxi_match:
        return daxi_match.group(1)
    xiaoxi_match = REG_XIAOXI.search(delivery_address)
    if xiaoxi_match:
        return xiaoxi_match.group(1)
    if "大西" in delivery_address:
        return "大西"
    elif "小西" in delivery_address:
        return "小西"
    else:
        return delivery_address

def get_address_base_sheet_name(delivery_address):
    for keyword, sheet_name in ADDRESS_SHEET_MAP.items():
        if keyword in delivery_address:
            return sheet_name
    if "农林" in delivery_address and "联建" not in delivery_address:
        return "东湖"
    return None

# ===================== Excel写入函数 =====================
def write_to_weekday_sheet(sheet, merged_ranges, order_info, meal_type):
    col_map = WEEKDAY_SHEET_COL_MAP[meal_type]
    row = get_first_empty_row(sheet, merged_ranges, start_col=col_map["单号"])
    meal_info = order_info["meal_info"]
    for key, col in col_map.items():
        cell_coord = f"{col}{row}"
        if cell_coord in merged_ranges:
            continue
        if key == "单号":
            sheet[cell_coord] = order_info["单号"]
        elif key == "姓名":
            sheet[cell_coord] = order_info["姓名"] or ""
        elif key == "地址":
            sheet[cell_coord] = order_info["处理后地址"]
        elif key == "电话":
            sheet[cell_coord] = order_info["电话"] or ""
        elif key == "经济/豪华":
            sheet[cell_coord] = meal_info.get("经济/豪华", "")
        elif key == "总餐次":
            sheet[cell_coord] = meal_info.get("总餐次", "")
    print(f"星期表[{sheet.title}] {meal_type} 第{row}行：{order_info['单号']}")
    return row

def write_to_address_meal_sheet(sheet, merged_ranges, order_info, meal_type):
    col_map = ADDRESS_MEAL_COL_MAP
    row = get_first_empty_row(sheet, merged_ranges, start_col=col_map["单号"])
    meal_info = order_info["meal_info"]
    weekday_fill = get_weekday_fill_value()
    for key in ["单号", "姓名", "地址", "电话"]:
        cell_coord = f"{col_map[key]}{row}"
        if cell_coord in merged_ranges:
            continue
        if key == "单号":
            sheet[cell_coord] = order_info["单号"]
        elif key == "姓名":
            sheet[cell_coord] = order_info["姓名"] or ""
        elif key == "地址":
            sheet[cell_coord] = order_info["处理后地址"]
        elif key == "电话":
            sheet[cell_coord] = order_info["电话"] or ""
    for weekday in WEEKDAY_INDEX_TO_NAME.values():
        cell_coord = f"{col_map[weekday]}{row}"
        if cell_coord in merged_ranges:
            continue
        sheet[cell_coord] = weekday_fill[weekday]
    for key in ["类型", "经济/豪华", "总餐次"]:
        cell_coord = f"{col_map[key]}{row}"
        if cell_coord in merged_ranges:
            continue
        if key == "类型":
            sheet[cell_coord] = meal_type
        elif key == "经济/豪华":
            sheet[cell_coord] = meal_info.get("经济/豪华", "")
        elif key == "总餐次":
            sheet[cell_coord] = meal_info.get("总餐次", "")
    print(f"地址表[{sheet.title}] {meal_type} 第{row}行：{order_info['单号']}")
    return row

def init_excel_sheets(wb, address_base_sheet, weekday_sheet_name):
    sheet_cache = {}
    if weekday_sheet_name not in wb.sheetnames:
        wb.create_sheet(title=weekday_sheet_name)
    weekday_sheet = wb[weekday_sheet_name]
    sheet_cache["weekday"] = {
        "sheet": weekday_sheet,
        "merged": cache_merged_ranges(weekday_sheet)
    }
    for meal_type in ["中餐", "晚餐"]:
        sheet_name = f"{address_base_sheet}{meal_type}"
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        sheet = wb[sheet_name]
        sheet_cache[meal_type] = {
            "sheet": sheet,
            "merged": cache_merged_ranges(sheet)
        }
    return sheet_cache

def process_order(order_info, lunch_infos, dinner_infos, wb, weekday_sheet_name, sheet_cache_map=None):
    address_base_sheet = order_info.get("地址基础表名")
    if not address_base_sheet:
        print(f"{order_info.get('单号')} 未匹配地址表，跳过")
        return
    if sheet_cache_map is None:
        sheet_cache_map = {}
    cache_key = (address_base_sheet, weekday_sheet_name)
    if cache_key in sheet_cache_map:
        sheet_cache = sheet_cache_map[cache_key]
    else:
        sheet_cache = init_excel_sheets(wb, address_base_sheet, weekday_sheet_name)
        sheet_cache_map[cache_key] = sheet_cache
    if lunch_infos:
        for meal_info in lunch_infos:
            order_info["meal_info"] = meal_info
            for _ in range(meal_info.get("count", 1)):
                write_to_weekday_sheet(
                    sheet_cache["weekday"]["sheet"],
                    sheet_cache["weekday"]["merged"],
                    order_info, "中餐"
                )
                write_to_address_meal_sheet(
                    sheet_cache["中餐"]["sheet"],
                    sheet_cache["中餐"]["merged"],
                    order_info, "中餐"
                )
    if dinner_infos:
        for meal_info in dinner_infos:
            order_info["meal_info"] = meal_info
            for _ in range(meal_info.get("count", 1)):
                write_to_weekday_sheet(
                    sheet_cache["weekday"]["sheet"],
                    sheet_cache["weekday"]["merged"],
                    order_info, "晚餐"
                )
                write_to_address_meal_sheet(
                    sheet_cache["晚餐"]["sheet"],
                    sheet_cache["晚餐"]["merged"],
                    order_info, "晚餐"
                )

def listen_keyboard_commands(stop_flag, restart_flag, reload_loop_flag):
    print("\n操作指令：- 终止 | + 重启 | ~ 重新加载当前循环")
    while True:
        try:
            cmd = input().strip()
            if cmd == "-":
                stop_flag.set()
                print("\n即将终止程序！")
                break
            elif cmd == "+":
                restart_flag.set()
                print("\n即将重启程序！")
                break
            elif cmd == "~":
                reload_loop_flag.set()
                print("\n即将重新加载当前循环！")
            else:
                print(f"\n无效指令「{cmd}」，请输入：- / + / ~")
        except:
            continue

# ===================== 主运行逻辑（核心修改部分已标注） =====================
def run_auto_operation():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel文件不存在：{EXCEL_PATH}")
        return
    
    stop_flag = threading.Event()
    restart_flag = threading.Event()
    reload_loop_flag = threading.Event()
    
    if os.environ.get('SKIP_COMMAND_LISTENER') == '1':
        print("跳过命令监听（测试模式：SKIP_COMMAND_LISTENER=1）")
    else:
        listen_thread = threading.Thread(
            target=listen_keyboard_commands,
            args=(stop_flag, restart_flag, reload_loop_flag),
            daemon=True
        )
        listen_thread.start()
    
    while not stop_flag.is_set():
        restart_flag.clear()
        reload_loop_flag.clear()
        
        wb = load_workbook(EXCEL_PATH)
        weekday_sheet_name = WEEKDAY_TO_SHEET[datetime.datetime.now().weekday()]

        sheets_to_clear = [
            weekday_sheet_name,
            "东湖中餐", "衣锦中餐", "医学院中餐",
            "东湖晚餐", "衣锦晚餐", "医学院晚餐"
        ]
        for sheet_name in sheets_to_clear:
            if sheet_name in wb.sheetnames:
                clear_sheet_from_row3(wb[sheet_name])
            else:
                print(f"表格[{sheet_name}]不存在，跳过清空")
        
        sheet_cache_map = {}
        
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=False,
                channel="msedge",
                args=["--window-size=1300,900"]
            )
            page = browser.new_page()
            
            try:
                if stop_flag.is_set():
                    break
                
                page.goto(TARGET_URL, timeout=ELEMENT_TIMEOUT, wait_until="networkidle")
                page.locator('input[placeholder="请输入手机号/账号"]').fill(PHONE_NUMBER)
                page.locator('input[placeholder="登录密码"]').fill(LOGIN_PASSWORD)
                page.locator("text=立即登录").click()
                page.wait_for_url("**/workbench/store", timeout=ELEMENT_TIMEOUT)
                print("登录成功")
                
                if stop_flag.is_set():
                    break
                
                page.locator('div.detail:has-text("门店地址")').dblclick()
                page.wait_for_url("**/home", timeout=ELEMENT_TIMEOUT)
                page.locator('div.navBarItem:has-text("订单")').click()
                page.wait_for_url("**/order/**", timeout=ELEMENT_TIMEOUT)
                page.locator('text=外送订单').click()
                page.wait_for_load_state("networkidle")
                print("进入外送订单列表")
                
                if stop_flag.is_set():
                    break
                
                total_wn = None
                while True:
                    total_wn_input = input("\n请输入待处理Wn订单总数（≥1）：").strip()
                    if total_wn_input.isdigit() and int(total_wn_input) >= 1:
                        total_wn = int(total_wn_input)
                        break
                    print("请输入≥1的纯数字！")
                
                current_page = 1
                wn_num = total_wn
                processed_count = 0
                while wn_num > 0 and not stop_flag.is_set():
                    if restart_flag.is_set():
                        print("\n执行重启指令，退出当前循环...")
                        break
                    if reload_loop_flag.is_set():
                        print(f"\n重新加载循环：剩余待处理 W{wn_num} - W1")
                        reload_loop_flag.clear()
                        current_page = 1
                        continue
                    
                    order_code = f"W{wn_num}"
                    processed_count = total_wn - wn_num + 1
                    print(f"\n===== 处理 [{processed_count}/{total_wn}]：{order_code} | 页码：{current_page} =====")
                    order_found = False
                    
                    while current_page <= MAX_PAGE_SEARCH and not stop_flag.is_set() and not restart_flag.is_set():
                        if not jump_to_page(page, current_page):
                            break
                        
                        # ================== 核心修改：搜索 + 重试 ==================
                        search_ok = False
                        for attempt in range(2):  # 最多尝试2次同一页
                            if attempt > 0:
                                page.wait_for_timeout(RETRY_WAIT)  # 等 300ms 再试
                            try:
                                # 精确匹配订单号
                                order_cell = page.locator(f'//td[normalize-space()="{order_code}"]').first
                                order_cell.wait_for(timeout=ORDER_SEARCH_TIMEOUT)
                                # 找到所在行，点击“详情”
                                detail_btn = order_cell.locator('xpath=ancestor::tr').locator('text=详情').first
                                detail_btn.click()
                                search_ok = True
                                break  # 成功，跳出重试循环
                            except TimeoutError:
                                if attempt == 0:
                                    print(f"  {order_code} 首次搜索超时，等待 {RETRY_WAIT}ms 重试...")
                                continue
                            except Exception:
                                raise  # 其他异常直接抛出
                        
                        if not search_ok:
                            # 两次都没找到，翻页
                            current_page += 1
                            print(f"{order_code} 第{current_page-1}页未找到，试第{current_page}页")
                            continue
                        # ============================================================
                        
                        page.wait_for_timeout(MEDIUM_WAIT)
                        order_found = True
                        
                        receiver_full = extract_text_after_label(page, "收货人")
                        receiver_name, receiver_phone = parse_receiver_info(receiver_full)
                        delivery_address = extract_text_after_label(page, "配送地址")
                        address_base_sheet = get_address_base_sheet_name(delivery_address)
                        
                        if address_base_sheet == "衣锦":
                            product_note = extract_product_note_text(page)
                            address_processed = get_yijin_address_from_product_note(product_note)
                        elif address_base_sheet == "医学院":
                            address_processed = delivery_address
                        elif address_base_sheet == "东湖":
                            address_processed = get_donghu_address_segment(delivery_address)
                        else:
                            address_processed = delivery_address
                        
                        lunch_infos = extract_meal_info(page, "午餐")
                        dinner_infos = extract_meal_info(page, "晚餐")
                        
                        order_info = {
                            "单号": order_code,
                            "姓名": receiver_name,
                            "电话": receiver_phone,
                            "处理后地址": address_processed,
                            "地址基础表名": address_base_sheet
                        }
                        print(f"订单信息：{order_info}")
                        
                        process_order(order_info, lunch_infos, dinner_infos, wb, weekday_sheet_name, sheet_cache_map)
                        
                        page.go_back()
                        page.wait_for_load_state("networkidle")
                        page.wait_for_timeout(SHORT_WAIT)  # 现在是 500ms，表格基本渲染完毕
                        break
                    
                    if not order_found:
                        print(f"{order_code} 未找到，跳过")
                    wn_num -= 1
                
                save_success = save_excel_with_retry(wb, EXCEL_PATH)
                browser.close()
                wb.close()
                
                if restart_flag.is_set():
                    print("\n重启程序中...")
                    continue
                if stop_flag.is_set():
                    print(f"\n程序已终止")
                    break
                
                if save_success:
                    print(f"\n所有{total_wn}个订单处理完成！")
                else:
                    print(f"\n订单处理完成，但Excel保存失败！")
                input("\n按任意键退出...")
                break
                
            except Exception as e:
                print(f"\n程序异常：{str(e)}")
                import traceback
                traceback.print_exc()
                save_excel_with_retry(wb, EXCEL_PATH)
                wb.close()
                browser.close()
                if restart_flag.is_set():
                    continue
                break

if __name__ == "__main__":
    print("依赖检查：")
    print("   1. pip install openpyxl playwright")
    print("   2. playwright install msedge")
    print(f"\n配置确认：Excel={EXCEL_PATH} | 账号={PHONE_NUMBER}")
    run_auto_operation()

