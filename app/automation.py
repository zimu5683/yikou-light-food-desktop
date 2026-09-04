"""Playwright order automation and Excel integration.

This module contains no credentials or machine-specific paths.  The GUI passes
an :class:`AppConfig`, a password and a cancellation event to ``run_job``.
Network selectors intentionally mirror the current admin site, while parsing
and Excel helpers remain usable in unit tests without Playwright installed.
"""
from __future__ import annotations

import datetime as _dt
import json
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any, Callable, Iterable

try:
    from .models import MealInfo, OrderInfo
    from .processing import (
        get_address_base_sheet_name,
        get_donghu_address_segment,
        get_yijin_address_from_product_note,
        parse_receiver_info,
    )
except ImportError:  # pragma: no cover - allows ``python app/automation.py``
    from models import MealInfo, OrderInfo
    from processing import (
        get_address_base_sheet_name,
        get_donghu_address_segment,
        get_yijin_address_from_product_note,
        parse_receiver_info,
    )

try:
    from .config import user_data_dir
    from .locators import DEFAULT_LOCATORS, load_locators
except ImportError:  # pragma: no cover - allows ``python app/automation.py``
    from config import user_data_dir
    from locators import DEFAULT_LOCATORS, load_locators

REG_MEAL_COUNT = re.compile(r"x\s*(\d+)", re.I)
REG_MEAL_SPLIT = re.compile(r"（午餐）|（晚餐）")
MAX_PAGE_SEARCH = 20
SHEET_MEAL_SUFFIX = {"午餐": "中餐", "晚餐": "晚餐"}
WEEKDAYS = ("周一", "周二", "周三", "周四", "周五", "周六", "周日")
HISTORICAL_SHEET_HEADERS = (
    "取单号", "姓名", "地址", "电话", *WEEKDAYS, "餐别", "经济/豪华", "总餐次",
)


def parse_target_date(value: object = None, *, today: _dt.date | None = None) -> _dt.date:
    """Parse a target date, defaulting to the local current date."""
    current = today or _dt.date.today()
    if value is None:
        return current
    if isinstance(value, _dt.datetime):
        result = value.date()
    elif isinstance(value, _dt.date):
        result = value
    else:
        text = str(value).strip()
        if not text:
            return current
        try:
            result = _dt.date.fromisoformat(text)
        except ValueError as exc:
            raise ValueError("目标日期格式必须为 YYYY-MM-DD") from exc
    if result > current:
        raise ValueError("目标日期不能晚于今天")
    return result


def parse_order_created_date(value: object) -> _dt.date | None:
    """Normalize common API date strings and Unix timestamps to a local date."""
    if value in (None, ""):
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    if isinstance(value, (int, float)):
        try:
            timestamp = float(value)
            if timestamp > 10_000_000_000:
                timestamp /= 1000
            return _dt.datetime.fromtimestamp(timestamp).date()
        except (OverflowError, OSError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        try:
            return parse_order_created_date(int(text))
        except ValueError:
            return None
    match = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        try:
            return _dt.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


class BrowserNotFoundError(RuntimeError):
    """Raised when no supported system browser is available."""

    def __init__(self, browsers: dict[str, str | None] | None = None) -> None:
        self.browsers = browsers or detect_browsers()
        super().__init__("未检测到可用浏览器，请安装 Microsoft Edge/Google Chrome，或安装 Playwright Chromium")


class LocatorError(RuntimeError):
    """Raised when a UI step matches none of its locator candidates.

    A screenshot, an HTML snapshot and the current URL are saved to the user
    log directory before this error is raised, so the site change can be
    fixed by editing the locator configuration instead of the code.
    """


def detect_browsers() -> dict[str, str | None]:
    """Find system browsers and an already-installed Playwright Chromium."""
    candidates: dict[str, list[Path]] = {"msedge": [], "chrome": [], "chromium": []}
    for name in candidates:
        found = shutil.which(name) or shutil.which(name + ".exe")
        if found:
            candidates[name].append(Path(found))
    if os.name == "nt":
        roots = [Path(os.environ.get(k, "")) for k in ("PROGRAMFILES", "PROGRAMFILES(X86)", "LOCALAPPDATA")]
        candidates["msedge"] += [root / "Microsoft/Edge/Application/msedge.exe" for root in roots if str(root)]
        candidates["chrome"] += [root / "Google/Chrome/Application/chrome.exe" for root in roots if str(root)]
    elif sys.platform == "darwin":
        candidates["msedge"] += _macos_browser_paths("msedge")
        candidates["chrome"] += _macos_browser_paths("chrome")
    else:
        candidates["msedge"] += _linux_browser_paths("msedge")
        candidates["chrome"] += _linux_browser_paths("chrome")
    chromium = _playwright_chromium_path()
    if chromium:
        candidates["chromium"].append(chromium)
    return {name: next((str(path) for path in paths if path.is_file()), None) for name, paths in candidates.items()}


def _macos_browser_paths(browser: str) -> list[Path]:
    app_roots = (Path("/Applications"), Path.home() / "Applications")
    app_name, executable = {
        "msedge": ("Microsoft Edge.app", "Microsoft Edge"),
        "chrome": ("Google Chrome.app", "Google Chrome"),
    }[browser]
    return [root / app_name / "Contents" / "MacOS" / executable for root in app_roots]


def _linux_browser_paths(browser: str) -> list[Path]:
    """Resolve distro browser executables whose names differ from ``msedge``/``chrome``."""
    commands, fixed = {
        "msedge": (("microsoft-edge", "microsoft-edge-stable"), ("/opt/microsoft/msedge/msedge",)),
        "chrome": (("google-chrome", "google-chrome-stable"), ("/opt/google/chrome/chrome",)),
    }[browser]
    paths = [Path(found) for name in commands if (found := shutil.which(name))]
    paths += [Path(target) for target in fixed]
    return paths


_CHROMIUM_PATH_CACHE: Path | None = None


def _playwright_chromium_path() -> Path | None:
    """Return Playwright's Chromium executable when the browser payload exists.

    Each lookup spawns Playwright's Node driver subprocess, so a positive
    result is cached for the lifetime of the process.  A failed lookup is not
    cached: after ``_install_chromium`` the next call must find the new
    payload without explicit invalidation.
    """
    global _CHROMIUM_PATH_CACHE
    if _CHROMIUM_PATH_CACHE is not None:
        return _CHROMIUM_PATH_CACHE
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as playwright:
            path = Path(playwright.chromium.executable_path)
        if path.is_file():
            _CHROMIUM_PATH_CACHE = path
            return path
    except Exception:
        pass
    return None


def _emit(callback: Callable[[str], Any] | None, message: str) -> None:
    if callback:
        callback(message)


def _base_url(config: Any) -> str:
    """Return the SPA base URL (scheme://host/path plus ``#`` for hash routing)."""
    url = str(getattr(config, "target_url", getattr(config, "url", "")) or "").strip()
    if not url:
        return ""
    if "#" in url:
        return url.split("#")[0].rstrip("/") + "/#"
    from urllib.parse import urlsplit
    parts = urlsplit(url)
    path = parts.path.rsplit("/", 1)[0].rstrip("/")
    return f"{parts.scheme}://{parts.netloc}{path}/"


def _build_locator(page: Any, candidate: dict[str, Any]) -> Any:
    """Build a Playwright locator for one candidate entry.

    The kind is detected from the keys present: ``css``, ``role`` (with
    optional ``name``/``name_re``), ``placeholder``, ``text`` or ``text_re``,
    and ``xpath`` (a full XPath expression such as ``//input``).
    ``has_text``/``has_text_re`` further filter the matched set.
    """
    if "css" in candidate:
        locator = page.locator(candidate["css"])
    elif "role" in candidate:
        kwargs: dict[str, Any] = {}
        if candidate.get("name_re"):
            kwargs["name"] = re.compile(candidate["name_re"])
        elif "name" in candidate:
            kwargs["name"] = candidate["name"]
        locator = page.get_by_role(candidate["role"], **kwargs)
    elif "placeholder" in candidate:
        locator = page.get_by_placeholder(candidate["placeholder"])
    elif "text" in candidate:
        locator = page.get_by_text(candidate["text"])
    elif "text_re" in candidate:
        locator = page.get_by_text(re.compile(candidate["text_re"]))
    elif "xpath" in candidate:
        locator = page.locator(candidate["xpath"])
    else:
        raise ValueError(f"无法识别的定位器候选: {candidate!r}")
    if candidate.get("has_text_re"):
        locator = locator.filter(has_text=re.compile(candidate["has_text_re"]))
    elif candidate.get("has_text"):
        locator = locator.filter(has_text=candidate["has_text"])
    return locator


def _find_by_candidates(page: Any, step: dict[str, Any], timeout: int) -> tuple[Any, int]:
    """Return ``(locator, candidate_index)`` for the first candidate that matches."""
    for index, candidate in enumerate(step.get("candidates", [])):
        try:
            locator = _build_locator(page, candidate)
            if locator.count() == 0:
                continue
            narrowed = locator.nth(int(candidate["index"])) if "index" in candidate else locator.first
            narrowed.wait_for(state="visible", timeout=min(max(timeout, 1), 3000))
            return narrowed, index
        except Exception:
            continue
    return None, -1


def _save_failure_snapshot(page: Any, step_name: str) -> Path:
    """Save screenshot, HTML and URL evidence for a failed UI step."""
    log_dir = user_data_dir() / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    safe = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", step_name)
    for extension, writer in (
        (".png", lambda path: page.screenshot(path=str(path), full_page=True)),
        (".html", lambda path: path.write_text(page.content(), encoding="utf-8")),
        (".txt", lambda path: path.write_text(page.url, encoding="utf-8")),
    ):
        try:
            writer(log_dir / f"{stamp}_{safe}{extension}")
        except Exception:
            pass
    return log_dir


def _locator_failure(page: Any, step_name: str) -> LocatorError:
    log_dir = _save_failure_snapshot(page, step_name)
    return LocatorError(
        f"页面定位失败：{step_name}。已保存页面截图、HTML 与当前网址到 {log_dir}，"
        f"可据此修改定位器配置（locators.json）后重试，无需更新程序。"
    )


def _locator_step(locators: dict[str, Any] | None, name: str) -> dict[str, Any]:
    """Return a step, falling back to the built-in default when missing."""
    return (locators or {}).get(name) or DEFAULT_LOCATORS.get(name) or {}


def _label_step(locators: dict[str, Any] | None, label: str) -> dict[str, Any]:
    user = ((locators or {}).get("labels") or {}).get(label)
    if user:
        return user
    default = (DEFAULT_LOCATORS.get("labels") or {}).get(label)
    return default or {"candidates": [{"text": label}]}


def _navigate(page: Any, step_name: str, locators: dict[str, Any] | None, base_url: str,
              timeout: int, callback: Callable[[str], Any] | None) -> None:
    """Open a page: direct URL first, then the ordered locator chain."""
    step = _locator_step(locators, step_name)
    wait_url = step.get("wait_url")
    if step.get("goto"):
        original_url = page.url
        url = str(step["goto"]).replace("{base}", base_url)
        try:
            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
            if wait_url:
                page.wait_for_url(wait_url, timeout=min(max(timeout, 1), 5000))
            _emit(callback, f"{step_name}：URL 直达成功")
            if step.get("confirm") == "table":
                _wait_for_order_table(page, timeout)
            return
        except Exception:
            try:
                if page.url != original_url:
                    page.goto(original_url, timeout=timeout, wait_until="domcontentloaded")
            except Exception:
                pass
    element, index = _find_by_candidates(page, step, timeout)
    if element is None:
        raise _locator_failure(page, step_name)
    try:
        if step.get("action") == "dblclick":
            element.dblclick(timeout=timeout)
        else:
            element.click(timeout=timeout)
        if step.get("wait_networkidle"):
            page.wait_for_load_state("networkidle", timeout=timeout)
        if wait_url:
            page.wait_for_url(wait_url, timeout=timeout)
        if step.get("confirm") == "table":
            _wait_for_order_table(page, timeout)
    except Exception:
        raise _locator_failure(page, step_name) from None
    _emit(callback, f"{step_name}：第 {index + 1} 个定位候选命中")


def parse_meal_rows(rows: Iterable[dict[str, str]], meal_type: str) -> list[MealInfo]:
    result: list[MealInfo] = []
    for row in rows:
        product = str(row.get("product", ""))
        quantity = str(row.get("qty", ""))
        segments = REG_MEAL_SPLIT.split(product)
        labels = REG_MEAL_SPLIT.findall(product)
        for index, segment in enumerate(segments[:-1]):
            current = "午餐" if labels[index] == "（午餐）" else "晚餐"
            if current != meal_type or not segment.strip():
                continue
            count_match = REG_MEAL_COUNT.search(quantity)
            result.append(MealInfo(
                total_meals=6 if "六餐" in segment else 1 if "单点" in segment else None,
                grade="经济" if "经济" in segment else "豪华" if "豪华" in segment else None,
                count=int(count_match.group(1)) if count_match else 1,
                meal_type=meal_type,
            ))
    return result


MEAL_ROWS_JS = """rows => rows.map(r => ({product:(r.querySelector('td:nth-child(1)')||{}).innerText||'', qty:(r.querySelector('td:nth-child(3)')||{}).innerText||''}))"""


def _meal_rows(page: Any, locators: dict[str, Any] | None) -> list[dict[str, str]]:
    """Collect meal rows using the first table candidate that yields data."""
    step = _locator_step(locators, "meal_table_row")
    for candidate in step.get("candidates", []):
        css = candidate.get("css")
        if not css:
            continue
        try:
            rows = page.eval_on_selector_all(css, MEAL_ROWS_JS)
            if rows:
                return rows or []
        except Exception:
            continue
    return []


def extract_meal_info(page: Any, meal_type: str, locators: dict[str, Any] | None = None) -> list[MealInfo]:
    return parse_meal_rows(_meal_rows(page, locators), meal_type)


def extract_product_note_text(page: Any, locators: dict[str, Any] | None = None) -> str:
    """Collect free-form notes rendered below product names."""
    step = _locator_step(locators, "meal_table_row")
    products: list[str] = []
    for candidate in step.get("candidates", []):
        css = candidate.get("css")
        if not css:
            continue
        try:
            products = page.eval_on_selector_all(
                css,
                """rows => rows.map(r => (r.querySelector('td:nth-child(1)') || {}).innerText || '')""",
            )
            if products:
                break
        except Exception:
            continue
    lines: list[str] = []
    for product in products or []:
        parts = [line.strip() for line in str(product).splitlines() if line.strip()]
        lines.extend(parts[1:])
    return " ".join(lines)


def ensure_browser(mode: str = "auto", allow_install: bool = True) -> str:
    """Return a usable browser, installing Playwright Chromium when needed."""
    mode = (mode or "auto").lower()
    found = detect_browsers()
    if mode in {"msedge", "edge"}:
        if found["msedge"]:
            return "msedge"
        raise BrowserNotFoundError(found)
    if mode in {"chrome", "google-chrome"}:
        if found["chrome"]:
            return "chrome"
        raise BrowserNotFoundError(found)
    if found["msedge"]:
        return "msedge"
    if found["chrome"]:
        return "chrome"
    if found.get("chromium"):
        return "chromium"
    if allow_install:
        try:
            _install_chromium()
        except Exception as exc:
            error = BrowserNotFoundError(found)
            error.__cause__ = exc
            raise error from exc
        found = detect_browsers()
        if found.get("chromium"):
            return "chromium"
    raise BrowserNotFoundError(found)


def _install_chromium() -> None:
    """Install Chromium using Playwright's bundled Node driver.

    Calling ``sys.executable -m playwright`` does not work from a PyInstaller
    executable because ``sys.executable`` points back to the application.  The
    driver and CLI are included by the PyInstaller Playwright hook, so invoking
    them directly works in both source and packaged builds.
    """
    from playwright._impl._driver import compute_driver_executable, get_driver_env

    driver, cli = compute_driver_executable()
    result = subprocess.run(
        [driver, cli, "install", "chromium"],
        env=get_driver_env(),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )
    if result.returncode != 0:
        # 输出可能很长（下载进度），只保留尾部关键错误信息用于排障。
        tail = (result.stdout or "").strip()[-500:]
        raise RuntimeError(f"Playwright Chromium 安装失败（退出码 {result.returncode}）：{tail}")


def _launch_browser(playwright: Any, mode: str, headless: bool) -> Any:
    # 任务启动不静默下载浏览器：打包版里下载没有任何进度提示，缺浏览器时
    # 报错交给 GUI 的「检查浏览器」按钮或 --install-browser 显式安装。
    allow_install = str(mode or "auto").lower() == "auto" and not getattr(sys, "frozen", False)
    preferred = ensure_browser(mode, allow_install=allow_install)
    kwargs = {"headless": headless, "args": ["--window-size=1300,900"]}
    executable = detect_browsers().get(preferred)
    if not executable:
        raise BrowserNotFoundError()
    return playwright.chromium.launch(executable_path=executable, **kwargs)


def _label(page: Any, label: str, timeout: int, locators: dict[str, Any] | None = None) -> str:
    """Read the value shown next to a detail label (e.g. 收货人 → 张三)."""
    element, _ = _find_by_candidates(page, _label_step(locators, label), timeout)
    if element is None:
        return ""
    try:
        matched = element.inner_text().strip()
        parent_text = element.locator("..").inner_text().strip()
        if matched and matched in parent_text:
            return parent_text.split(matched)[-1].lstrip("：:").strip()
        return parent_text.split(label)[-1].lstrip("：:").strip()
    except Exception:
        return ""


def _historical_sheet_name(target_date: _dt.date) -> str:
    return f"{target_date.year}年{target_date.month}月{target_date.day}日 {WEEKDAYS[target_date.weekday()]}"


def _write_historical_order(wb: Any, order: OrderInfo, meal: MealInfo, meal_type: str,
                            target_date: _dt.date) -> None:
    """Append an old order to its dedicated date sheet instead of today's plan."""
    sheet_name = _historical_sheet_name(target_date)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    if sheet.max_row == 1 and all(sheet.cell(1, index).value in (None, "")
                              for index in range(1, len(HISTORICAL_SHEET_HEADERS) + 1)):
        for index, title in enumerate(HISTORICAL_SHEET_HEADERS, 1):
            sheet.cell(1, index).value = title
    row = max(2, sheet.max_row + 1)
    while sheet.cell(row, 1).value not in (None, ""):
        row += 1
    weekday = WEEKDAYS[target_date.weekday()]
    values = [
        order.order_no, order.name, order.address, order.phone,
        *[1 if day == weekday else "" for day in WEEKDAYS],
        meal_type, meal.grade or "", meal.total_meals or "",
    ]
    for index, value in enumerate(values, 1):
        sheet.cell(row, index).value = value


def _write_order(wb: Any, order: OrderInfo, meal: MealInfo, meal_type: str,
                 target_date: _dt.date | None = None, today: _dt.date | None = None) -> None:
    """Write today's plan normally, or archive a selected historical date."""
    target_date = target_date or _dt.date.today()
    today = today or _dt.date.today()
    if target_date < today:
        _write_historical_order(wb, order, meal, meal_type, target_date)
        return
    base = order.address_base_sheet
    if not base:
        return
    weekday = WEEKDAYS[(_dt.datetime.now().weekday() + 1) % 7]
    weekday_sheet = wb[weekday] if weekday in wb.sheetnames else wb.create_sheet(weekday)
    target_name = f"{base}{SHEET_MEAL_SUFFIX.get(meal_type, meal_type)}"
    target = wb[target_name] if target_name in wb.sheetnames else wb.create_sheet(target_name)
    columns = ("A", "B", "C", "D", "E", "F") if meal_type == "午餐" else ("G", "H", "I", "J", "K", "L")
    row = max(3, weekday_sheet.max_row + 1)
    while weekday_sheet[f"{columns[0]}{row}"].value not in (None, ""):
        row += 1
    values = (order.order_no, order.name, order.address, order.phone, meal.grade or "", meal.total_meals or "")
    for col, value in zip(columns, values):
        weekday_sheet[f"{col}{row}"] = value
    row2 = max(3, target.max_row + 1)
    while target[f"A{row2}"].value not in (None, ""):
        row2 += 1
    vals = [order.order_no, order.name, order.address, order.phone] + [1 if d == weekday else "" for d in WEEKDAYS] + [meal_type, meal.grade or "", meal.total_meals or ""]
    for idx, value in enumerate(vals, 1):
        target.cell(row2, idx).value = value


def _load_order_workbook(excel_path: Path, loader: Callable[..., Any] | None = None) -> Any:
    if loader is None:
        from openpyxl import load_workbook
        loader = load_workbook
    return loader(excel_path, keep_vba=excel_path.suffix.lower() == ".xlsm")


def _detail_api_json(page: Any, detail_url: str) -> dict[str, Any] | None:
    """在已登录会话内直接请求详情接口，绕过前端渲染层。

    站点前端存在渲染缺陷（接口数据到达却不渲染、路由随机弹回首页），
    而 /channel/order/{id} 接口稳定返回完整 JSON。鉴权走 Bearer token
    （localStorage.layout_token）与 uniacid 头，不走 Cookie。
    """
    match = re.search(r"[?&]id=(\d+)", detail_url or "")
    if not match:
        return None
    store_match = re.search(r"[?&]storeId=(\d+)", detail_url or "")
    store_id = store_match.group(1) if store_match else ""
    script = (
        "async ({oid, sid}) => {"
        "const r = await fetch(`/channel/order/${oid}?storeId=${sid}`, {"
        "credentials: 'include',"
        "headers: {'authorization': `Bearer ${localStorage.layout_token}`,"
        "'uniacid': localStorage.layout_uniacid || ''}});"
        "return await r.text();"
        "}"
    )
    try:
        raw = page.evaluate(script, {"oid": match.group(1), "sid": store_id})
        payload = json.loads(raw)
    except Exception:
        return None
    if not isinstance(payload, dict) or not isinstance(payload.get("data"), dict):
        return None
    return payload["data"]


def _order_from_api(page: Any, code: str, detail_url: str) -> OrderInfo | None:
    """从详情接口 JSON 构造订单；接口不可用或字段全空时返回 None。

    字段映射与页面渲染一致：address.contact=收货人、address.mobile=电话、
    address.address+description=配送地址、goods[].name/num=商品与数量
    （名称自带（午餐）/（晚餐）标签）、attrData.matal=商品备注。
    """
    data = _detail_api_json(page, detail_url)
    if not data:
        return None
    address_info = data.get("address") or {}
    address = " ".join(str(address_info.get(k) or "").strip() for k in ("address", "description")).strip()
    name = str(address_info.get("contact") or "").strip()
    phone = str(address_info.get("mobile") or data.get("mobile") or "").strip()
    if not name:
        name = str((data.get("user") or {}).get("nickname") or "").strip()
    rows = [{"product": str(g.get("name") or ""), "qty": f"x {g.get('num') or 1}"}
            for g in (data.get("goods") or []) if g.get("name")]
    notes: list[str] = []
    for goods in data.get("goods") or []:
        attr = goods.get("attrData") or {}
        if attr.get("matal"):
            notes.append(str(attr["matal"]))
        for material in attr.get("material") or []:
            if isinstance(material, dict) and material.get("name"):
                notes.append(str(material["name"]))
    base = get_address_base_sheet_name(address)
    if base == "东湖":
        address = get_donghu_address_segment(address)
    elif base == "衣锦":
        address = get_yijin_address_from_product_note(" ".join(notes))
    metadata = {
        "order_id": re.search(r"[?&]id=(\d+)", detail_url or "").group(1)
        if re.search(r"[?&]id=(\d+)", detail_url or "") else "",
        "created_at": next((data.get(key) for key in
                             ("created_at", "createdAt", "create_time", "createTime", "order_time", "orderTime")
                             if data.get(key) not in (None, "")), None),
    }
    candidate = OrderInfo(code, name, phone, address, base, delivery_address=address, metadata=metadata)
    candidate.lunch = parse_meal_rows(rows, "午餐")
    candidate.dinner = parse_meal_rows(rows, "晚餐")
    if not (name or phone or address or candidate.lunch or candidate.dinner):
        return None
    return candidate


def _read_order(page: Any, code: str, timeout: int, locators: dict[str, Any] | None,
                callback: Callable[[str], Any] | None, detail_url: str = "") -> OrderInfo | None:
    """读取详情页并构造 OrderInfo；返回 None 表示两次尝试都读到空数据。

    站点 SPA 会在任意时刻自发重新引导回 #/home（channelLogin/loadMenus
    重跑后 router 重置），读取途中随时可能被弹走。一旦发现读到的数据
    全空、或读完时已不在详情路由，就用详情地址重新直达再读一遍。
    detail_url 必须由调用方在详情验证通过后传入——不能用当前地址，
    否则页面已被弹走时会“重读”到首页。
    """
    if not detail_url:
        try:
            url = page.url
        except Exception:
            url = ""
        detail_url = url if _is_order_detail_url(url) else ""
    for attempt in range(2):
        if attempt:
            if not detail_url:
                return None
            try:
                if _is_order_detail_url(getattr(page, "url", "")):
                    _emit(callback, f"{code} 详情数据为空，重新加载详情后重读")
                else:
                    _emit(callback, f"{code} 读取期间页面被弹走，重新直达详情重读")
                page.goto(detail_url, timeout=timeout, wait_until="domcontentloaded")
            except Exception as exc:
                _emit(callback, f"{code} 重读详情失败：{exc}")
                return None
            try:
                _wait_for_detail(page, timeout, callback)
            except Exception:
                pass
        name, phone = _read_contact(page, timeout, locators)
        address = _read_address(page, timeout, locators)
        base = get_address_base_sheet_name(address)
        if base == "东湖":
            address = get_donghu_address_segment(address)
        elif base == "衣锦":
            address = get_yijin_address_from_product_note(extract_product_note_text(page, locators))
        candidate = OrderInfo(code, name, phone, address, base, delivery_address=address)
        for typ, attr in (("午餐", "lunch"), ("晚餐", "dinner")):
            setattr(candidate, attr, extract_meal_info(page, typ, locators))
        if (name or phone or address or candidate.lunch or candidate.dinner) \
                and _is_order_detail_url(getattr(page, "url", "")):
            try:
                api_data = _detail_api_json(page, detail_url)
            except Exception:
                api_data = None
            if api_data:
                candidate.metadata.update({
                    "order_id": re.search(r"[?&]id=(\d+)", detail_url or "").group(1)
                    if re.search(r"[?&]id=(\d+)", detail_url or "") else "",
                    "created_at": next((api_data.get(key) for key in
                                         ("created_at", "createdAt", "create_time", "createTime", "order_time", "orderTime")
                                         if api_data.get(key) not in (None, "")), None),
                })
            return candidate
        # DOM 读空通常不是页面改版，而是站点前端缺陷：接口数据到了却不渲染，
        # 或路由随机弹回首页。数据在 /channel/order/{id} 里，直接取 JSON。
        if not attempt and detail_url:
            api_candidate = _order_from_api(page, code, detail_url)
            if api_candidate is not None:
                _emit(callback, f"{code} 页面未渲染出详情数据，已改从详情接口读取")
                return api_candidate
    return None


def run_job(config: Any, order_count: int, stop_event: Any, progress_callback: Callable[[str], Any] | None = None, password: str | None = None,
            order_decision_callback: Callable[[str, str], str] | None = None,
            save_decision_callback: Callable[[str], str] | None = None,
            locators: dict[str, Any] | None = None,
            target_date: object = None) -> dict[str, int]:
    """Process the newest W orders and append their meals to the workbook."""
    configured_excel = getattr(config, "excel_path", None)
    if not configured_excel:
        raise FileNotFoundError("尚未选择 Excel 文件")
    excel_path = Path(configured_excel)
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("仅支持 .xlsx 和 .xlsm Excel 文件")
    if order_count < 1:
        raise ValueError("order_count 必须大于等于 1")
    selected_date = parse_target_date(
        target_date if target_date is not None else getattr(config, "order_date", "")
    )
    today = _dt.date.today()
    if password is None:
        password = getattr(config, "password", "")
    if locators is None:
        locators = load_locators()
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError("缺少 Playwright，请先安装 requirements.txt") from exc
    backup_dir = excel_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    shutil.copy2(excel_path, backup_dir / f"{excel_path.stem}_{stamp}{excel_path.suffix}")
    # Preserve embedded VBA when the user explicitly selects an .xlsm file.
    wb = _load_order_workbook(excel_path)
    processed = 0
    found = 0
    timeout = int(getattr(config, "element_timeout_ms", 8000))
    try:
        with sync_playwright() as playwright:
            browser = _launch_browser(playwright, getattr(config, "browser_mode", "auto"), bool(getattr(config, "headless", False)))
            page = browser.new_page()
            try:
                _emit(progress_callback, "正在登录...")
                page.goto(getattr(config, "target_url", getattr(config, "url", "")), timeout=timeout, wait_until="networkidle")
                base_url = _base_url(config)

                account_input, _ = _find_by_candidates(page, _locator_step(locators, "login_account_input"), timeout)
                if account_input is None:
                    raise _locator_failure(page, "账号输入框")
                account_input.fill(getattr(config, "phone_number", getattr(config, "phone", "")))

                password_input, _ = _find_by_candidates(page, _locator_step(locators, "login_password_input"), timeout)
                if password_input is None:
                    raise _locator_failure(page, "密码输入框")
                password_input.fill(password or "")

                submit, _ = _find_by_candidates(page, _locator_step(locators, "login_submit"), timeout)
                if submit is None:
                    raise _locator_failure(page, "登录按钮")
                submit.click(timeout=timeout)
                login_step = _locator_step(locators, "登录成功")
                page.wait_for_url(str(login_step.get("wait_url") or "**/workbench/store"), timeout=timeout)

                # Navigation and locator recovery are implementation details;
                # only the order-level status is shown in the normal log.
                _navigate(page, "门店地址", locators, base_url, timeout, None)
                _navigate(page, "订单菜单", locators, base_url, timeout, None)
                # 站点数据层不稳定（Tab 点击后表格可能长时间空白），最多重试三轮导航。
                last_table_error: Exception | None = None
                for _ in range(3):
                    try:
                        _navigate(page, "外送订单", locators, base_url, timeout, None)
                        page.wait_for_load_state("networkidle")
                        _wait_for_order_table(page, timeout)
                        last_table_error = None
                        break
                    except Exception as exc:
                        last_table_error = exc
                        _emit(progress_callback, "订单列表暂未就绪，正在重试")
                if last_table_error is not None:
                    raise last_table_error
                _emit(progress_callback, "登录成功，开始处理订单")
                list_url = page.url
                for number in range(order_count, 0, -1):
                    if stop_event.is_set():
                        break
                    code = f"W{number}"
                    order: OrderInfo | None = None
                    occurrence = 0
                    while not stop_event.is_set():
                        try:
                            _ensure_waimai_tab(page, locators, base_url, timeout, None, list_url)
                            try:
                                cell = _find_order_cell(page, code, config, None, occurrence=occurrence)
                            except LookupError:
                                if occurrence:
                                    _emit(progress_callback, f"{code} 没有目标日期订单，未写入 Excel")
                                    break
                                raise
                            try:
                                list_url = page.url
                            except Exception:
                                list_url = ""
                            _open_detail(page, code, cell, config, None, timeout, list_url, occurrence)
                            _wait_for_detail(page, timeout, None)
                            try:
                                detail_url = page.url
                            except Exception:
                                detail_url = ""
                            candidate = _read_order(page, code, timeout, locators, None, detail_url)
                            if candidate is None:
                                _save_failure_snapshot(page, f"详情空数据_{code}")
                                raise LookupError(
                                    f"订单 {code} 详情页未读到姓名/电话/地址/餐品，疑似页面改版或详情被弹回首页，"
                                    f"已放弃本次写入（快照见 logs 目录）")
                            created_date = parse_order_created_date(candidate.metadata.get("created_at"))
                            if created_date is None:
                                _emit(progress_callback, f"{code} 缺少下单日期，已跳过")
                                occurrence += 1
                                _back_to_order_list(page, timeout, list_url, reason="跳过", callback=None)
                                continue
                            if created_date != selected_date:
                                _emit(progress_callback, f"{code} 跳过 {created_date.isoformat()} 订单（目标 {selected_date.isoformat()}）")
                                occurrence += 1
                                _back_to_order_list(page, timeout, list_url, reason="跳过", callback=None)
                                continue

                            # Do not mutate the workbook until the browser has
                            # safely returned to the order list.  A retry can
                            # therefore never duplicate partially written rows.
                            if not _back_to_order_list(page, timeout, list_url, reason="返回", callback=None):
                                raise LookupError(f"订单 {code} 已读取但无法返回订单列表")
                            order = candidate
                            break
                        except Exception as exc:
                            _recover_order_table(page, timeout, list_url)
                            _emit(progress_callback, f"{code} 处理失败：{exc}")
                            decision = "skip"
                            if order_decision_callback:
                                decision = order_decision_callback(code, str(exc)).lower()
                            if decision == "retry":
                                _emit(progress_callback, f"重试 {code}")
                                continue
                            if decision == "stop":
                                _emit(progress_callback, f"{code} 已选择停止，本轮结束")
                                stop_event.set()
                                break
                            _emit(progress_callback, f"{code} 未找到，跳过")
                            break

                    if stop_event.is_set():
                        break
                    processed += 1
                    if order is None:
                        _emit(progress_callback, "-------")
                        continue
                    for typ, meals in (("午餐", order.lunch), ("晚餐", order.dinner)):
                        for meal in meals:
                            for _ in range(max(1, meal.count)):
                                _write_order(wb, order, meal, typ, target_date=selected_date, today=today)
                    meal_text = _format_order_meals(order)
                    _emit(progress_callback, _format_order_summary(order, meal_text))
                    found += 1
                    _emit(progress_callback, "-------")
            finally:
                browser.close()
        _save_workbook_with_retry(wb, excel_path, save_decision_callback)
    finally:
        wb.close()
    _emit(progress_callback, f"处理完成：找到 {found}/{processed} 个订单")
    return {"processed": processed, "found": found}


def _save_workbook_with_retry(workbook: Any, excel_path: Path,
                              decision_callback: Callable[[str], str] | None = None) -> None:
    """Save an Excel workbook, allowing the user to close a locked file and retry."""
    while True:
        try:
            workbook.save(str(excel_path))
            return
        except PermissionError as exc:
            if decision_callback is None:
                raise
            decision = decision_callback(str(exc)).strip().lower()
            if decision not in {"retry", "重试", "再次保存"}:
                raise PermissionError(f"已取消保存 Excel 文件：{excel_path}") from exc


def _format_order_meals(order: OrderInfo) -> str:
    parts: list[str] = []
    for meal_type, meals in (("午餐", order.lunch), ("晚餐", order.dinner)):
        for meal in meals:
            grade = meal.grade or "未标注"
            total = f"{meal.total_meals}餐" if meal.total_meals else "餐品"
            parts.append(f"{meal_type}{grade}{total} x{meal.count}")
    return "、".join(parts) if parts else "未识别"


def _format_order_summary(order: OrderInfo, meal_text: str | None = None) -> str:
    """Render one compact, user-facing line for a successfully read order."""
    meals = meal_text if meal_text is not None else _format_order_meals(order)
    return "｜".join((
        order.order_no or "未知订单",
        order.name or "未填写",
        order.phone or "未填写",
        order.address or "未填写",
        meals,
    ))


def _wait_for_order_table(page: Any, timeout: int) -> None:
    """Wait for the SPA to render at least one order row."""
    page.wait_for_load_state("domcontentloaded", timeout=timeout)
    page.locator(".el-table__body-wrapper tbody tr:visible, table tbody tr:visible").first.wait_for(
        state="visible", timeout=timeout
    )


def _recover_order_table(page: Any, timeout: int, list_url: str = "") -> bool:
    """Best-effort recovery after a detail-page or rendering failure.

    注意：只在确认当前不在订单列表页时才后退/回跳，避免把正常的列表页
    越退越远（历史退穿是 W8 之后 W7 找不到的主因）。
    """
    try:
        _wait_for_order_table(page, min(timeout, 1200))
        return True
    except Exception:
        pass
    return _back_to_order_list(page, timeout, list_url, reason="恢复")


def _is_order_detail_url(url: str) -> bool:
    return "/order/detail" in (url or "")


def _waimai_tab_active(page: Any) -> bool:
    """外送订单 Tab 是否处于选中状态（只读检查，绝不点击）。"""
    for build in (
        lambda: page.get_by_role("tab", name=re.compile(r"外送订单")).first,
        lambda: page.locator(".el-tabs__item", has_text=re.compile(r"外送订单")).first,
    ):
        try:
            tab = build()
        except Exception:
            continue
        try:
            if tab.count() == 0:
                continue
            cls = (tab.get_attribute("class") or "").lower()
            if "active" in cls or "selected" in cls or "current" in cls:
                return True
            if tab.get_attribute("aria-selected") == "true":
                return True
        except Exception:
            continue
        return False
    return False


def _ensure_waimai_tab(page: Any, locators: dict[str, Any] | None, base_url: str,
                        timeout: int, callback: Callable[[str], Any] | None,
                        list_url: str = "") -> None:
    """每次搜索前确认停在 外送订单 Tab 上：只验证，不盲点。

    列表页是 自提订单/外送订单/店内订单… 多 Tab，任何重载或路由 remount
    都会把它重置回默认的 自提订单（W 单全在外送订单下），而表格等待在
    错误 Tab 上也能通过。之前“每单前重点一次 Tab”会在已选中的 Tab 上
    触发表格重刷，正好撞上紧随其后的行定位——这就是首单 W8 也失败的
    regression。MISS 时才点，命中则零操作。

    站点 SPA 会不定期把页面弹回 #/home（列表默认 Tab 和详情页都可能中招），
    此时 Tab 根本不存在，直接找 Tab 必然报“页面定位失败：外送订单”。
    所以恢复必须先 goto 回记录的列表地址，再点 Tab。
    """
    try:
        url = page.url
    except Exception:
        url = ""
    on_list = "takeOutList" in (url or "")
    if not on_list and list_url:
        _emit(callback, f"页面被弹离订单列表（当前 {url or '未知'}），直达回列表页")
        try:
            page.goto(list_url, timeout=timeout, wait_until="domcontentloaded")
        except Exception as exc:
            _emit(callback, f"回列表页直达失败：{exc}")
    elif not on_list and "/order" not in (url or ""):
        _emit(callback, f"当前不在订单列表页（{url or '未知'}），回到外送订单")
    try:
        if on_list and _waimai_tab_active(page):
            return
    except Exception:
        pass
    _emit(callback, "外送订单 Tab 未确认选中，重新点击")
    _navigate(page, "外送订单", locators, base_url, timeout, callback)


def _open_detail(page: Any, code: str, cell: Any, config: Any,
                  callback: Callable[[str], Any] | None, timeout: int,
                  list_url: str = "", occurrence: int = 0) -> None:
    """点击详情并验证地址真的进了 detail 路由；没点开就明错，绝不执行返回。

    点击落在表格重渲染间隙时会丢失（地址不变、无新历史），此时若执行
    返回就会一路退到 #/home。这就是“点了外送订单却跳首页”的机制。
    若页面已被 SPA 弹回 #/home（Tab 不在列表页），先直达回列表再重找该行。
    """
    try:
        before = page.url
    except Exception:
        before = ""
    now = before
    last_err: Exception | None = None
    for i in range(2):
        try:
            btn = cell.locator("xpath=ancestor::tr").locator("text=详情").first
            btn.click(timeout=timeout)
            page.wait_for_url("**/order/detail**", timeout=timeout)
            page.wait_for_load_state("domcontentloaded", timeout=timeout)
            return
        except Exception as exc:
            last_err = exc
            try:
                now = page.url
            except Exception:
                now = before
            if _is_order_detail_url(now):
                return
            if i == 0:
                _emit(callback, f"{code} 详情首次点击无跳转（地址无变化），重找该行再点一次")
                try:
                    if list_url and "takeOutList" not in (now or ""):
                        page.goto(list_url, timeout=timeout, wait_until="domcontentloaded")
                    cell = _find_order_cell(page, code, config, callback, occurrence=occurrence)
                except Exception:
                    pass
                continue
    raise LookupError(f"订单 {code} 详情未点开（点击后地址仍为：{now or before or '未知'}），已放弃，绝不执行返回") from last_err


def _wait_for_detail(page: Any, timeout: int, callback: Callable[[str], Any] | None = None) -> None:
    """点详情后确认真的进了详情路由，且关键字段已渲染。

    站点 SPA 可能在进入 detail 路由后立刻把页面弹回 #/home（内容变成
    首页仪表盘），此时直接读取必然全空。发现弹回（或字段迟迟不渲染）
    就用详情地址重新直达一次——直达加载出的详情页是完整的。
    """
    page.wait_for_url("**/order/detail**", timeout=timeout)
    page.wait_for_load_state("domcontentloaded", timeout=timeout)
    try:
        detail_url = page.url
    except Exception:
        detail_url = ""
    pattern = re.compile(r"收货人|收件人|下单人|联系电话")
    for attempt in range(2):
        try:
            page.get_by_text(pattern).first.wait_for(
                state="visible", timeout=min(max(timeout, 1), 5000)
            )
            if _is_order_detail_url(getattr(page, "url", "")):
                return
        except Exception:
            pass
        if not detail_url or attempt > 0:
            break
        try:
            here = page.url
        except Exception:
            here = "未知"
        if _is_order_detail_url(here):
            _emit(callback, "详情页字段未渲染，用详情地址重新直达")
        else:
            _emit(callback, f"详情页被弹回 {here}，用详情地址重新直达")
        try:
            page.goto(detail_url, timeout=timeout, wait_until="domcontentloaded")
        except Exception as exc:
            _emit(callback, f"重新直达详情失败：{exc}")
            return
    _emit(callback, "详情页字段等待超时，尝试直接读取")


def _read_contact(page: Any, timeout: int, locators: dict[str, Any] | None) -> tuple[str, str]:
    """按新旧两套标签读取姓名电话：收货人 -> 下单人 -> 联系电话。"""
    for label in ("收货人", "下单人"):
        value = _label(page, label, timeout, locators)
        name, phone = parse_receiver_info(value)
        if name or phone:
            return name, phone
    phone = _label(page, "联系电话", timeout, locators).strip()
    name = _label(page, "下单人", timeout, locators).strip()
    if not name:
        name, phone2 = parse_receiver_info(phone)
        phone = phone2 or phone
    return name, phone


def _read_address(page: Any, timeout: int, locators: dict[str, Any] | None) -> str:
    for label in ("配送地址", "收货地址", "送餐地址", "用户备注"):
        value = _label(page, label, timeout, locators).strip()
        if value:
            return value
    return ""


def _back_to_order_list(page: Any, timeout: int, list_url: str, reason: str = "返回",
                        callback: Callable[[str], Any] | None = None) -> bool:
    """从详情页回到订单列表：彻底不用浏览器历史，一律用记录的列表地址直达。

    历史后退在 SPA 里不可靠（详情没点开时一次 go_back 就会弹回 #/home）；
    直达会重置 Tab，调用方下次搜索前会重新确认外送订单 Tab。
    """
    if not _is_order_detail_url(getattr(page, "url", "")):
        # 根本没进详情，绝不能碰历史，直接确认表格还在即可。
        try:
            _wait_for_order_table(page, min(timeout, 2000))
            return True
        except Exception:
            pass
    if list_url:
        try:
            page.goto(list_url, timeout=timeout, wait_until="domcontentloaded")
            _wait_for_order_table(page, timeout)
            _emit(callback, f"{reason}列表成功（直达）：{list_url}")
            return True
        except Exception as exc:
            _emit(callback, f"{reason}列表直达失败：{exc}")
    try:
        _wait_for_order_table(page, timeout)
        return True
    except Exception:
        return False


def _find_order_cell(page: Any, code: str, config: Any, callback: Callable[[str], Any] | None,
                     occurrence: int = 0) -> Any:
    """Find an exact order number, retrying renders and traversing pages."""
    timeout = max(1000, int(getattr(config, "order_search_timeout_ms", 8000)))
    pause = max(200, int(getattr(config, "retry_wait_ms", 1000)))
    attempts = max(1, int(getattr(config, "order_search_attempts", 3)))
    max_pages = max(1, int(getattr(config, "max_page_search", MAX_PAGE_SEARCH)))

    # Searches should always start from page one; after returning from details
    # the SPA may otherwise leave the pagination on the previous order's page.
    first_page = page.locator('.el-pagination li.number').filter(has_text=re.compile(r'^\s*1\s*$')).first
    try:
        if first_page.is_visible() and "active" not in (first_page.get_attribute("class") or ""):
            first_page.click()
            page.wait_for_timeout(pause)
    except Exception:
        pass

    last_error: Exception | None = None
    for page_number in range(1, max_pages + 1):
        matches = page.locator('.el-table__body-wrapper tbody tr:visible td, table tbody tr:visible td').filter(
            has_text=re.compile(rf'^\s*{re.escape(code)}\s*$')
        )
        for attempt in range(1, attempts + 1):
            try:
                count = matches.count()
                if count > occurrence:
                    cell = matches.first if occurrence == 0 else matches.nth(occurrence)
                    cell.wait_for(state="visible", timeout=timeout)
                    return cell
                raise LookupError(f"订单 {code} 不在第 {page_number} 页")
            except Exception as exc:
                last_error = exc
                if attempt < attempts:
                    _emit(callback, f"{code} 页面仍在刷新，{pause / 1000:g} 秒后重试 ({attempt}/{attempts - 1})")
                    page.wait_for_timeout(pause)

        if page_number >= max_pages:
            break
        # ``occurrence`` counts matching rows across all previous pages.
        # Subtract this page before looking at the next page.
        occurrence = max(0, occurrence - matches.count())
        next_button = page.locator('.el-pagination button.btn-next, button.btn-next').first
        try:
            disabled = next_button.is_disabled() or next_button.get_attribute("disabled") is not None
            disabled = disabled or "disabled" in (next_button.get_attribute("class") or "").lower()
            if disabled:
                break
            _emit(callback, f"{code} 当前页未找到，继续搜索第 {page_number + 1} 页")
            next_button.click()
            page.wait_for_timeout(pause)
            _wait_for_order_table(page, timeout)
        except Exception as exc:
            last_error = exc
            break
    raise last_error or TimeoutError(f"订单 {code} 未找到")


__all__ = ["run_job", "ensure_browser", "detect_browsers", "BrowserNotFoundError", "LocatorError", "parse_receiver_info", "parse_meal_rows", "extract_meal_info", "extract_product_note_text", "get_yijin_address_from_product_note", "get_address_base_sheet_name", "get_donghu_address_segment"]
