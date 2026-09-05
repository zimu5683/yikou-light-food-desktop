"""闪时送（sss）平台批量下单自动化（接口模式）。

本模块与管理后台订单处理（:mod:`app.automation`）方向相反：从独立的
《闪时送.xlsx》读取订单（午餐/晚餐两表），再在闪时送平台逐单创建预约单。

界面只负责登录：程序自动切换“账户密码登录”并填写账号密码，用户只需输入
图形验证码（输完第 4 位自动点登录，验证码错误会提示重输）。下单本身不再
操作页面表单——登录成功后通过页面内 ``fetch`` 直调后端接口（``token`` 头
取自 ``localStorage.tokenObj``），与订单处理侧的
:func:`app.automation.run_job` 同一套思路，不受页面弹回与渲染缺陷影响。

干跑模式：配置 ``sss_dry_run = true`` 时只组装并打印下单报文，不真实提交。
凭据与路径由 GUI 通过 :class:`AppConfig` 传入，不写入源码。
"""
from __future__ import annotations

import datetime as _dt
import json
import time
from pathlib import Path
from typing import Any, Callable

try:
    from .automation import (
        BrowserNotFoundError,
        LocatorError,
        _emit,
        _launch_browser,
    )
    from .locators import SSS_LOCATORS, load_sss_locators
except ImportError:  # pragma: no cover - allows ``python app/sss.py``
    from automation import (
        BrowserNotFoundError,
        LocatorError,
        _emit,
        _launch_browser,
    )
    from locators import SSS_LOCATORS, load_sss_locators

DEFAULT_SHEETS = ("午餐", "晚餐")
LUNCH_TIME = "11:00:00"
DINNER_TIME = "17:00:00"
DEFAULT_SSS_URL = "https://sssplusnew.zhuopaikeji.com/takeout"

# 下单相关接口（2026-09-05 实测抓包确认，详见 .zcode/sss_api_recon.md）。
_CREATE_ORDER_PATH = "/consumer/order/one-touch-send/create-order-from-client"
_STORE_LIST_PATH = "/consumer/customer/store/queryStoreAddresses?pageNo=1&pageSize=40"
_FREQUENT_ADDR_PATH = "/consumer/customer/customerAddress/queryFrequentAddressByCustomer"

_SSS_FETCH_JS = """
async ({method, path, body}) => {
  let token = null;
  try {
    const t = JSON.parse(localStorage.getItem('tokenObj') || 'null');
    if (t && t.expirationTime > Date.now()) token = t.token;
  } catch (e) {}
  const headers = {'Content-Type': 'application/json'};
  if (token) headers['token'] = token;
  const r = await fetch(path, {method, headers,
      body: body === null || body === undefined ? undefined : JSON.stringify(body)});
  return JSON.stringify({http: r.status, text: await r.text()});
}
"""


def _clean(value: Any) -> Any:
    return value.strip() if isinstance(value, str) else value


def load_sss_orders(excel_path: str | Path, sheets=DEFAULT_SHEETS) -> dict[str, list[dict[str, Any]]]:
    """读取《闪时送.xlsx》的订单（A=姓名 B=门牌号 C=电话 D=送达时间）。

    每张工作表从第 3 行开始，遇到 A/B/C 三列均为空的行即终止。返回
    ``{工作表名: [订单字典, ...]}``，每个订单含 ``row`` 与四列原始值。
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    try:
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            orders: list[dict[str, Any]] = []
            row_num = 3
            while True:
                name = _clean(ws[f"A{row_num}"].value)
                door = _clean(ws[f"B{row_num}"].value)
                phone = _clean(ws[f"C{row_num}"].value)
                delivery_time = ws[f"D{row_num}"].value
                if not any([name, door, phone]):
                    break
                orders.append({
                    "row": row_num,
                    "name": name,
                    "door": door,
                    "phone": phone,
                    "delivery_time": delivery_time,
                })
                row_num += 1
            result[sheet_name] = orders
        return result
    finally:
        wb.close()


def compute_delivery_time(is_dinner: bool, now: _dt.datetime | None = None) -> str:
    """按原脚本规则计算送达时间：午餐 11:00 / 晚餐 17:00，16 点后顺延次日。"""
    now = now or _dt.datetime.now()
    target_date = now
    hour = now.hour
    if 20 <= hour < 24 or 16 <= hour < 20:  # 原脚本的等价写法：16~23 点顺延次日
        target_date = now + _dt.timedelta(days=1)
    time_str = DINNER_TIME if is_dinner else LUNCH_TIME
    return target_date.strftime("%Y-%m-%d ") + time_str


def _resolve(locators: dict[str, Any] | None, name: str) -> dict[str, Any]:
    """Return a step dict, falling back to the built-in 闪时送 default."""
    return (locators or {}).get(name) or SSS_LOCATORS.get(name) or {}


def _substitute_tokens(step: dict[str, Any], **tokens: str) -> dict[str, Any]:
    """Deep-copy a step and replace ``{token}`` placeholders in candidate strings."""
    copied = json.loads(json.dumps(step))

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            for key, value in list(node.items()):
                if isinstance(value, str):
                    for token, replacement in tokens.items():
                        value = value.replace("{" + token + "}", replacement)
                    node[key] = value
                else:
                    walk(value)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    walk(copied)
    return copied


def _pick(record: dict[str, Any], keys: tuple[str, ...]) -> Any:
    """按优先级从记录里取第一个非空字段。"""
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return None


def _result_records(payload: dict[str, Any]) -> list:
    """从 {result: {records: [...]}} 或 {result: [...]} 里取记录列表。"""
    result = payload.get("result")
    if isinstance(result, list):
        return result
    if isinstance(result, dict):
        records = result.get("records") or result.get("list")
        if isinstance(records, list):
            return records
    return []


def _match_record(records: Any, keyword: str, what: str) -> dict[str, Any]:
    """在列表里按关键词匹配（任意字段包含即可）；唯一记录时直接采用。"""
    if not isinstance(records, list) or not records:
        raise LookupError(f"{what}列表为空或格式异常")
    for record in records:
        if keyword and keyword in json.dumps(record, ensure_ascii=False):
            return record
    if len(records) == 1:
        return records[0]
    raise LookupError(
        f"{what}列表里没有匹配「{keyword}」的记录，共 {len(records)} 条："
        + json.dumps(records, ensure_ascii=False)[:600])


def _sss_api(page: Any, method: str, path: str, body: Any = None) -> tuple[int, dict[str, Any]]:
    """在登录页会话内调闪时送接口（token 头自动取自 localStorage.tokenObj）。"""
    raw = page.evaluate(_SSS_FETCH_JS, {"method": method, "path": path, "body": body})
    envelope = json.loads(raw)
    try:
        payload = json.loads(envelope.get("text") or "{}")
    except ValueError:
        payload = {}
    return int(envelope.get("http") or 0), payload


def _ensure_logged_in(page: Any, account: str, password: str, stop_event: Any,
                      callback: Callable[[str], Any] | None) -> None:
    """确保已登录：已登录直接返回；否则自动填写并等用户输验证码后自动点登录。

    验证码输错时页面会刷新出新验证码，循环等待重输；登录成功的标志是
    「创建订单」按钮出现（登录后进入的是订单页）。
    """
    if page.locator("text=创建订单").count():
        return
    page.wait_for_selector("text=账户密码登录", state="visible", timeout=30000)
    page.get_by_text("账户密码登录").click()
    page.wait_for_selector("input#account", state="visible", timeout=15000)
    page.fill("input#account", account)
    page.fill("input#password", password)
    _emit(callback, ">>> 请在闪时送窗口输入图形验证码（输完自动点登录；"
                    "窗口已最小化时请先点任务栏还原）<<<")

    deadline = time.time() + 300
    while time.time() < deadline:
        if stop_event.is_set():
            raise RuntimeError("已停止")
        code_val = page.evaluate(
            "() => (document.querySelector('input#code')||{}).value || ''")
        if len(code_val.strip()) >= 4:
            page.locator("button", has_text="登录").first.click()
            page.wait_for_timeout(2500)
            if page.locator("text=创建订单").count():
                _emit(callback, "登录成功")
                return
            # 验证码错误：清空输入框等待重输（否则残留旧值会触发误重试）
            try:
                page.fill("input#code", "")
            except Exception:
                pass
            _emit(callback, "验证码不正确或登录未成功，请按窗口里的新验证码重新输入")
        page.wait_for_timeout(400)
    raise TimeoutError("登录超时：未完成验证码输入")


def _minimize_window(browser: Any, callback: Callable[[str], Any] | None) -> None:
    """Best-effort 最小化浏览器窗口（CDP），失败不阻断流程。"""
    try:
        session = browser.new_browser_cdp_session()
        window_id = session.send("Browser.getWindowForTarget")["windowId"]
        session.send("Browser.setWindowBounds",
                     {"windowId": window_id, "bounds": {"windowState": "minimized"}})
        _emit(callback, "浏览器已最小化（任务栏可见，需要输验证码时再还原）")
    except Exception:
        pass


def build_order_payload(order: dict[str, Any], is_dinner: bool, store_id: int,
                        address: dict[str, Any], goods_name: str,
                        now: _dt.datetime | None = None) -> dict[str, Any]:
    """组装 create-order-from-client 的请求体（字段为 2026-09-05 抓包确认）。"""
    return {
        "expectedDeliveryTime": compute_delivery_time(is_dinner, now),
        "goodsDetail": [{"goodsName": goods_name, "goodsNum": 1}],
        "orderType": 2,  # 预约单
        "receiveName": str(order.get("name") or ""),
        "receivePhone": str(order.get("phone") or ""),
        "storeId": store_id,
        "receiveAddress": {
            "lnt": address["lnt"],
            "lat": address["lat"],
            "areaCode": address["areaCode"],
            "addressDetail": address["addressDetail"],
            "doorNum": str(order.get("door") or ""),
        },
    }


def _address_from_record(record: dict[str, Any]) -> dict[str, Any]:
    """从常用地址记录提取下单需要的坐标与地址字段。

    站点前端的映射（chunk 反解）：记录的 ``markLnglat`` 子对象携带
    ``longitude/latitude/adcode/address``，顶层字段仅作兜底。
    """
    mark = record.get("markLnglat") if isinstance(record.get("markLnglat"), dict) else {}

    def pick(src: dict[str, Any], keys: tuple[str, ...]) -> Any:
        for key in keys:
            value = src.get(key)
            if value not in (None, ""):
                return value
        return None

    lnt = pick(mark, ("longitude", "lng", "lnt", "lon")) or _pick(record, ("longitude", "lnt", "lng"))
    lat = pick(mark, ("latitude", "lat")) or _pick(record, ("latitude", "lat"))
    area = pick(mark, ("adcode", "areaCode")) or _pick(record, ("areaCode", "code"))
    detail = pick(mark, ("address", "addressDetail")) or _pick(record, ("position", "address", "addressDetail"))
    if lnt in (None, "") or lat in (None, ""):
        raise LookupError("常用地址记录缺少经纬度字段："
                          + json.dumps(record, ensure_ascii=False)[:400])
    return {
        "lnt": float(lnt), "lat": float(lat),
        "areaCode": str(area or ""), "addressDetail": str(detail or ""),
    }


def run_sss_job(config: Any, stop_event: Any,
                progress_callback: Callable[[str], Any] | None = None,
                password: str | None = None,
                decision_callback: Callable[[str, str], str] | None = None,
                locators: dict[str, Any] | None = None) -> dict[str, int]:
    """读取《闪时送.xlsx》并通过接口逐单创建预约单。

    ``decision_callback(identifier, error)`` 返回 ``retry``/``skip``/``stop``，
    用于单个订单创建失败时的交互决策（与订单处理的 order_decision 一致）。
    ``config.sss_dry_run`` 为真时只组装并打印报文，不真实提交。
    """
    configured_excel = getattr(config, "sss_excel_path", None)
    if not configured_excel:
        raise FileNotFoundError("尚未选择闪时送 Excel 文件")
    excel_path = Path(configured_excel)
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("仅支持 .xlsx 和 .xlsm Excel 文件")
    if password is None:
        password = ""
    if locators is None:
        locators = load_sss_locators()

    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError("缺少 Playwright，请先安装 requirements.txt") from exc

    account = str(getattr(config, "sss_account", "") or "")
    if not account:
        raise ValueError("尚未填写闪时送账号")
    dry_run = bool(getattr(config, "sss_dry_run", False))
    store_name = str(getattr(config, "sss_store_name", "") or "一口轻食")
    common_address = str(getattr(config, "sss_common_address", "") or "")
    goods_name = str(getattr(config, "sss_product_name", "") or "轻食")

    orders_by_sheet = load_sss_orders(excel_path)
    total = sum(len(orders) for orders in orders_by_sheet.values())
    if total == 0:
        _emit(progress_callback, "闪时送 Excel 中没有任何订单")
        return {"processed": 0, "created": 0}

    timeout = int(getattr(config, "element_timeout_ms", 8000))
    url = str(getattr(config, "sss_url", "") or "").strip() or DEFAULT_SSS_URL
    processed = 0
    created = 0

    with sync_playwright() as playwright:
        browser = _launch_browser(
            playwright,
            getattr(config, "browser_mode", "auto"),
            bool(getattr(config, "headless", False)),
        )
        page = browser.new_page()
        try:
            _emit(progress_callback, "正在打开闪时送…")
            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
            _ensure_logged_in(page, account, password, stop_event, progress_callback)
            _minimize_window(browser, progress_callback)

            # 登录会话就绪：读取门店与常用地址，之后每单只调下单接口。
            _emit(progress_callback, "读取门店与常用地址…")
            _, store_payload = _sss_api(page, "GET", _STORE_LIST_PATH)
            store_record = _match_record(_result_records(store_payload),
                                         store_name, "门店")
            store_id = _pick(store_record, ("id", "storeId", "storeNo"))
            if store_id in (None, ""):
                raise LookupError("门店记录缺少 id："
                                  + json.dumps(store_record, ensure_ascii=False)[:300])

            _, addr_payload = _sss_api(page, "GET", _FREQUENT_ADDR_PATH)
            address_record = _match_record(_result_records(addr_payload),
                                           common_address, "常用地址")
            _emit(progress_callback,
                  f"常用地址记录：{json.dumps(address_record, ensure_ascii=False)[:400]}")
            address = _address_from_record(address_record)
            _emit(progress_callback,
                  f"门店「{store_name}」与常用地址「{common_address}」就绪"
                  f"（{address['addressDetail']} @ {address['lnt']},{address['lat']}）")
            if dry_run:
                _emit(progress_callback, "【干跑模式】只组装报文，不真实提交")

            for sheet_name in DEFAULT_SHEETS:
                orders = orders_by_sheet.get(sheet_name, [])
                if not orders:
                    continue
                is_dinner = sheet_name == "晚餐"
                _emit(progress_callback, f"开始处理【{sheet_name}】表，共 {len(orders)} 单")
                for order in orders:
                    if stop_event.is_set():
                        break
                    processed += 1
                    identifier = f"第 {order['row']} 行 {order.get('name') or '未填写'}"
                    payload = build_order_payload(order, is_dinner, int(store_id),
                                                  address, goods_name)
                    if dry_run:
                        _emit(progress_callback,
                              f"【干跑】{identifier} 报文：{json.dumps(payload, ensure_ascii=False)}")
                        created += 1
                        continue
                    _emit(progress_callback, f"正在创建【{sheet_name}】{identifier}")
                    try:
                        http, resp = _submit_order(page, payload)
                        if http == 401 or resp.get("code") == 401:
                            # token 服务端过期（约 30-60 分钟）：重新登录后重试一次
                            _emit(progress_callback, "登录态过期，重新登录…")
                            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
                            _ensure_logged_in(page, account, password, stop_event,
                                              progress_callback)
                            _minimize_window(browser, progress_callback)
                            http, resp = _submit_order(page, payload)
                        if not resp.get("success"):
                            raise LookupError(
                                resp.get("message") or json.dumps(resp, ensure_ascii=False)[:200])
                        created += 1
                        _emit(progress_callback,
                              f"订单创建成功：{identifier}（预约 {payload['expectedDeliveryTime']}）")
                    except Exception as exc:
                        _emit(progress_callback, f"订单创建失败：{identifier}：{exc}")
                        if decision_callback is not None:
                            decision = decision_callback(identifier, str(exc)).lower()
                            if decision == "retry":
                                _emit(progress_callback, f"重试 {identifier}")
                                try:
                                    http, resp = _submit_order(page, payload)
                                    if not resp.get("success"):
                                        raise LookupError(
                                            resp.get("message") or "创建失败")
                                    created += 1
                                    _emit(progress_callback, f"重试成功：{identifier}")
                                except Exception as exc2:
                                    _emit(progress_callback, f"重试仍失败：{identifier}：{exc2}")
                            elif decision == "stop":
                                stop_event.set()
                                break
                if stop_event.is_set():
                    break
        finally:
            browser.close()

    mode = "（干跑）" if dry_run else ""
    _emit(progress_callback, f"闪时送下单完成{mode}：成功 {created}/{processed}")
    return {"processed": processed, "created": created}


def _submit_order(page: Any, payload: dict[str, Any]) -> tuple[int, dict[str, Any]]:
    http, resp = _sss_api(page, "POST", _CREATE_ORDER_PATH, payload)
    return http, resp


__all__ = ["run_sss_job", "load_sss_orders", "compute_delivery_time",
           "build_order_payload", "SSS_LOCATORS",
           "BrowserNotFoundError", "LocatorError"]
