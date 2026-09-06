"""Excel 模板生成，供桥接层（app/bridge.py）与未来的 UI 复用。

从旧 Tkinter 界面（app/gui.py）原样抽出，保持「第 2 行占位、数据从第 3 行
开始」的读写约定：订单处理按需建表，缺哪张表程序会自动创建；闪时送固定
《午餐/晚餐》两表。
"""
from __future__ import annotations

from pathlib import Path


def write_order_template(dest: Path) -> None:
    """生成排单模板：周一~周日 + 六个常用食堂表，全部写好表头。"""
    from openpyxl import Workbook

    wb = Workbook()
    weekday_headers = ["单号", "姓名", "地址", "电话", "经济/豪华", "总餐次"]
    weekdays = ("周一", "周二", "周三", "周四", "周五", "周六", "周日")
    first = True
    for name in (*weekdays, "衣锦中餐", "衣锦晚餐", "医学院中餐", "医学院晚餐", "东湖中餐", "东湖晚餐"):
        ws = wb.active if first else wb.create_sheet(title=name)
        if first:
            ws.title = name
            first = False
        ws.append(weekday_headers)
        ws.append([])  # 第 2 行为占位行，数据从第 3 行开始（与程序读写规则一致）
    wb.save(str(dest))
    wb.close()


def write_sss_template(dest: Path) -> None:
    """生成闪时送下单用的《午餐/晚餐》两表模板，含示例行。"""
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for name in ("午餐", "晚餐"):
        ws = wb.active if first else wb.create_sheet(title=name)
        if first:
            ws.title = name
            first = False
        ws["A1"] = "姓名"
        ws["B1"] = "门牌号"
        ws["C1"] = "电话"
        ws["D1"] = "送达时间"
        # 第 2 行占位，第 3 行起填数据；示例行方便照格式填写。
        ws["A2"] = ""
        ws["A3"] = "张三"
        ws["B3"] = "1栋101"
        ws["C3"] = "13800000000"
        ws["D3"] = "11:00" if name == "午餐" else "17:00"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 12
    wb.save(str(dest))
    wb.close()
