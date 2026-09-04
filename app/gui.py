"""Tkinter user interface for the order processor."""
from __future__ import annotations

import datetime as _dt
import queue
import os
import sys
import threading
import tkinter as tk
import webbrowser
from tkinter import filedialog, font as tkfont, messagebox, ttk

from .automation import BrowserNotFoundError, ensure_browser, parse_target_date, run_job
from .config import AppConfig, clamp_split_ratio
from .credentials import (delete_password, delete_sss_password, get_password,
                          get_sss_password, set_password, set_sss_password)
from .sss import run_sss_job
from .design_system import (FONT_FAMILY, FONT_MONO, PillButton, ResponsiveSplitPane,
                            RoundedCard, ScrollableRoundedCard, StatusBadge, TOKENS, FormField,
                            apply_tk_scaling, form_layout_mode, layout_mode)
from . import __version__
from .updater import ReleaseInfo, UpdateError, check_for_update, download_and_install


# Tk 在 Linux/X11 下对 ``("Excel 文件", "*.xlsx *.xlsm")`` 这种单字符串多后缀写法
# 支持不可靠（空格会被当成文件名的一部分，导致过滤后看不到任何文件）。
# 这里拆成逐后缀 + 元组形式，Windows/macOS/Linux 三端都可正确过滤。
# 注意：Linux 下“所有文件”必须用 ``*`` 而非 ``*.*``（后者只匹配带点的文件名）。
EXCEL_FILETYPES: tuple[tuple[str, str | tuple[str, ...]], ...] = (
    ("Excel 工作簿 (*.xlsx)", "*.xlsx"),
    ("Excel 启用宏的工作簿 (*.xlsm)", "*.xlsm"),
    ("Excel 文件 (*.xlsx *.xlsm)", ("*.xlsx", "*.xlsm")),
    ("所有文件 (*)", "*"),
)
EXCEL_EXTS = {".xlsx", ".xlsm"}


class _DatePickerPopup(tk.Toplevel):
    """A single-month, dependency-free calendar for one target date."""

    def __init__(self, owner: "App", anchor: tk.Widget, selected: _dt.date) -> None:
        super().__init__(owner)
        self.owner = owner
        self.anchor = anchor
        self.pending = selected
        self.today = _dt.date.today()
        self.month = selected.replace(day=1)
        self.configure(bg=TOKENS.surface, bd=1, highlightthickness=1, highlightbackground=TOKENS.border)
        self.overrideredirect(True)
        self.transient(owner)
        self._render()
        self.update_idletasks()
        x = max(0, min(anchor.winfo_rootx(), owner.winfo_screenwidth() - self.winfo_reqwidth() - 8))
        y = min(anchor.winfo_rooty() + anchor.winfo_height() + 2,
                owner.winfo_screenheight() - self.winfo_reqheight() - 8)
        self.geometry(f"+{x}+{y}")
        self.grab_set()
        self.bind("<Escape>", lambda _event: self.destroy())

    def _render(self) -> None:
        for child in self.winfo_children():
            child.destroy()
        header = tk.Frame(self, bg=TOKENS.surface)
        header.pack(fill="x", padx=12, pady=(10, 6))
        tk.Button(header, text="<", command=lambda: self._change_month(-1), relief="flat", bd=0,
                  bg=TOKENS.surface_muted, fg=TOKENS.primary, activebackground=TOKENS.green_light,
                  font=(FONT_FAMILY, 9, "bold"), padx=8, pady=3).pack(side="left")
        tk.Label(header, text=f"{self.month.year} 年 {self.month.month} 月", bg=TOKENS.surface, fg=TOKENS.text,
                 font=(FONT_FAMILY, 11, "bold")).pack(side="left", expand=True)
        next_button = tk.Button(header, text=">", command=lambda: self._change_month(1), relief="flat", bd=0,
                                bg=TOKENS.surface_muted, fg=TOKENS.primary, activebackground=TOKENS.green_light,
                                font=(FONT_FAMILY, 9, "bold"), padx=8, pady=3)
        if self._add_months(self.month, 1) > self.today.replace(day=1):
            next_button.configure(state="disabled", fg=TOKENS.text_soft)
        next_button.pack(side="right")

        calendar = tk.Frame(self, bg=TOKENS.surface)
        calendar.pack(fill="both", padx=12, pady=(0, 10))
        self._render_month(calendar, self.month)

        footer = tk.Frame(self, bg=TOKENS.surface, highlightthickness=1, highlightbackground=TOKENS.border_soft)
        footer.pack(fill="x")
        tk.Button(footer, text="清空", command=self._clear, relief="flat", bd=0,
                  bg=TOKENS.surface, fg=TOKENS.text_soft, activebackground=TOKENS.surface_muted,
                  font=(FONT_FAMILY, 9), padx=14, pady=8).pack(side="left", padx=5)
        tk.Button(footer, text="确定", command=self._confirm, relief="flat", bd=0,
                  bg=TOKENS.accent, fg=TOKENS.text_on_dark, activebackground=TOKENS.uplift,
                  activeforeground=TOKENS.text_on_dark, font=(FONT_FAMILY, 9, "bold"), padx=16, pady=8).pack(side="right", padx=5)

    @staticmethod
    def _add_months(value: _dt.date, delta: int) -> _dt.date:
        index = value.year * 12 + value.month - 1 + delta
        return _dt.date(index // 12, index % 12 + 1, 1)

    def _render_month(self, master: tk.Misc, month: _dt.date) -> None:
        panel = tk.Frame(master, bg=TOKENS.surface)
        panel.pack()
        for index, day_name in enumerate(("日", "一", "二", "三", "四", "五", "六")):
            tk.Label(panel, text=day_name, width=3, bg=TOKENS.surface, fg=TOKENS.text_soft,
                     font=(FONT_FAMILY, 9, "bold")).grid(row=0, column=index, padx=2, pady=(0, 3))
        first_cell = (month.weekday() + 1) % 7
        start = month - _dt.timedelta(days=first_cell)
        for index in range(42):
            value = start + _dt.timedelta(days=index)
            row, cell_column = divmod(index, 7)
            in_month = value.month == month.month
            selectable = in_month and value <= self.today
            button = tk.Button(panel, text=str(value.day), width=3, relief="flat", bd=0,
                               command=lambda day=value: self._select(day), font=(FONT_FAMILY, 9),
                               bg=TOKENS.accent if value == self.pending else TOKENS.surface,
                               fg=TOKENS.text_on_dark if value == self.pending else (TOKENS.text if selectable else TOKENS.text_soft),
                               activebackground=TOKENS.uplift, activeforeground=TOKENS.text_on_dark)
            if not selectable:
                button.configure(state="disabled", bg=TOKENS.surface, disabledforeground=TOKENS.text_soft)
            button.grid(row=row + 1, column=cell_column, padx=2, pady=1)

    def _change_month(self, delta: int) -> None:
        self.month = self._add_months(self.month, delta)
        self._render()

    def _select(self, value: _dt.date) -> None:
        self.pending = value
        self._render()

    def _clear(self) -> None:
        self.pending = None
        self._render()

    def _confirm(self) -> None:
        if self.pending is None:
            self.owner._clear_order_date()
        else:
            self.owner._set_order_date(self.pending)
        self.destroy()


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        apply_tk_scaling(self)
        try:
            self._dpi_value = float(self.winfo_fpixels("1i"))
        except (tk.TclError, ValueError):
            self._dpi_value = 72.0
        self.title("一口轻食 - 订单处理")
        self.geometry("1080x720")
        self.minsize(820, 620)
        self.events: queue.Queue[tuple[str, object]] = queue.Queue()
        self.stop_event = threading.Event()
        self.worker: threading.Thread | None = None
        self._closing = False
        self._log_lines: list[str] = []
        self._search_var = tk.StringVar()
        self._saved_config = AppConfig.load()
        self._split_ratio = clamp_split_ratio(self._saved_config.split_ratio)
        self._resize_job: str | None = None
        self._config_layout_job: str | None = None
        self._config_compact: bool | None = None
        self.configure(bg=TOKENS.canvas)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._build_widgets()
        self._load_saved_values()
        self.after(100, self._drain_events)
        self.after(700, self.check_for_updates)

    def _build_widgets(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TProgressbar", troughcolor=TOKENS.ceramic, background=TOKENS.accent,
                        bordercolor=TOKENS.ceramic, lightcolor=TOKENS.accent, darkcolor=TOKENS.accent)

        outer = tk.Frame(self, bg=TOKENS.canvas)
        outer.pack(fill="both", expand=True)
        outer.rowconfigure(1, weight=1)
        outer.columnconfigure(0, weight=1)

        # 104px assumes Segoe UI metrics; Linux fallback fonts are taller, so
        # size the header from real font metrics or the title gets clipped.
        title_font = tkfont.Font(family=FONT_FAMILY, size=24, weight="bold")
        subtitle_font = tkfont.Font(family=FONT_FAMILY, size=10)
        header_height = max(104, title_font.metrics("linespace") + TOKENS.space_1
                            + subtitle_font.metrics("linespace") + TOKENS.space_4 * 2)
        header = tk.Frame(outer, bg=TOKENS.house, height=header_height)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)
        header.columnconfigure(0, weight=1)
        header_inner = tk.Frame(header, bg=TOKENS.house)
        header_inner.grid(row=0, column=0, sticky="nsew", padx=TOKENS.space_6, pady=TOKENS.space_4)
        header_inner.columnconfigure(0, weight=1)
        tk.Label(header_inner, text="一口轻食", bg=TOKENS.house, fg=TOKENS.text_on_dark,
                 font=(FONT_FAMILY, 24, "bold"), anchor="w").grid(row=0, column=0, sticky="w")
        tk.Label(header_inner, text="订单自动处理中心  ·  配置任务并实时查看运行状态",
                 bg=TOKENS.house, fg=TOKENS.text_on_dark_soft, font=(FONT_FAMILY, 10),
                 anchor="w").grid(row=1, column=0, sticky="w", pady=(TOKENS.space_1, 0))
        tk.Label(header_inner, text=f"版本 {__version__}", bg=TOKENS.uplift, fg=TOKENS.text_on_dark,
                 font=(FONT_FAMILY, 9, "bold"), padx=12, pady=5).grid(row=0, column=1, rowspan=2, sticky="e")

        self.content = tk.Frame(outer, bg=TOKENS.canvas, padx=TOKENS.space_6, pady=TOKENS.space_5)
        self.content.grid(row=1, column=0, sticky="nsew")
        self.content.rowconfigure(0, weight=1)
        self.content.columnconfigure(0, weight=1)

        self.split_pane = ResponsiveSplitPane(
            self.content, ratio=self._split_ratio,
            on_ratio_committed=self._on_split_ratio_committed,
        )
        self.split_pane.grid(row=0, column=0, sticky="nsew")
        self.config_card = ScrollableRoundedCard(self.split_pane.left_host,
                                                 parent_bg=TOKENS.canvas, padding=TOKENS.space_4)
        self.log_card = RoundedCard(self.split_pane.right_host,
                                    parent_bg=TOKENS.canvas, padding=TOKENS.space_4)
        self.config_card.pack(fill="both", expand=True)
        self.log_card.pack(fill="both", expand=True)
        self._build_config_card()
        self._build_log_card()
        self._layout_mode = ""
        self.bind("<Configure>", self._on_resize)
        self.bind("<Configure>", self._refresh_dpi_scaling, add="+")
        self.after_idle(self._on_resize)
        # Geometry negotiation can finish after the first idle callback on
        # high-DPI Windows; run one more pass once the window is mapped.
        self.after(80, self._on_resize)

    def _refresh_dpi_scaling(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        """Re-read monitor DPI when a window is moved between displays."""
        try:
            dpi = float(self.winfo_fpixels("1i"))
        except (tk.TclError, ValueError):
            return
        if abs(dpi - self._dpi_value) >= 1.0:
            self._dpi_value = dpi
            apply_tk_scaling(self)

    def _build_config_card(self) -> None:
        body = self.config_card.body
        body.columnconfigure(0, weight=1)
        body.bind("<Configure>", self._on_config_body_configure, add="+")
        tk.Label(body, text="任务配置", bg=TOKENS.surface, fg=TOKENS.primary,
                 font=(FONT_FAMILY, 20, "bold"), anchor="w").grid(row=0, column=0, sticky="w")
        tk.Label(body, text="选择要执行的任务类型，并准备好相应资料。", bg=TOKENS.surface,
                 fg=TOKENS.text_soft, font=(FONT_FAMILY, 10), anchor="w").grid(row=1, column=0, sticky="w", pady=(TOKENS.space_1, TOKENS.space_4))

        self._build_mode_switch(body)

        self.mode_host = tk.Frame(body, bg=TOKENS.surface)
        self.mode_host.grid(row=3, column=0, sticky="ew")
        self.mode_host.columnconfigure(0, weight=1)
        self.order_form = tk.Frame(self.mode_host, bg=TOKENS.surface)
        self.sss_form = tk.Frame(self.mode_host, bg=TOKENS.surface)
        self._build_order_form()
        self._build_sss_form()
        self._switch_mode("order")
        self._apply_config_layout(520)

    def _build_mode_switch(self, body: tk.Misc) -> None:
        switch = tk.Frame(body, bg=TOKENS.surface)
        switch.grid(row=2, column=0, sticky="ew", pady=(0, TOKENS.space_3))
        switch.columnconfigure(0, weight=1)
        switch.columnconfigure(1, weight=1)
        self.mode_buttons = {
            "order": PillButton(switch, "订单处理", lambda: self._switch_mode("order"), variant="primary"),
            "sss": PillButton(switch, "闪时送下单", lambda: self._switch_mode("sss"), variant="outline"),
        }
        self.mode_buttons["order"].grid(row=0, column=0, sticky="ew", padx=(0, TOKENS.space_1))
        self.mode_buttons["sss"].grid(row=0, column=1, sticky="ew", padx=(TOKENS.space_1, 0))

    def _switch_mode(self, mode: str) -> None:
        self._mode = mode
        self.order_form.grid_remove()
        self.sss_form.grid_remove()
        target = self.order_form if mode == "order" else self.sss_form
        target.grid(row=0, column=0, sticky="ew")
        for key, button in self.mode_buttons.items():
            button.set_variant("primary" if key == mode else "outline")
        self.start_button = self.order_start_button if mode == "order" else self.sss_start_button
        self.stop_button = self.order_stop_button if mode == "order" else self.sss_stop_button

    def _build_order_form(self) -> None:
        body = self.order_form
        body.columnconfigure(0, weight=1)

        fields = [("管理网址", "url", "例如：https://example.com/admin"),
                  ("手机号 / 账号", "phone", "用于登录管理后台"),
                  ("登录密码", "password", "密码仅保存在系统凭据管理器中")]
        self.vars: dict[str, tk.StringVar] = {key: tk.StringVar() for _, key, _ in fields if key != "password"}
        self.password = tk.StringVar()
        self.form_fields: dict[str, FormField] = {}
        for row, (label, key, helper) in enumerate(fields, start=0):
            variable = self.password if key == "password" else self.vars[key]
            field = FormField(body, label, variable, show="*" if key == "password" else "", helper=helper)
            field.grid(row=row, column=0, sticky="ew", pady=(0, TOKENS.space_3))
            self.form_fields[key] = field

        excel_row = tk.Frame(body, bg=TOKENS.surface)
        self.excel_row = excel_row
        excel_row.grid(row=3, column=0, sticky="ew", pady=(0, TOKENS.space_3))
        excel_row.columnconfigure(0, weight=1)
        self.vars["excel"] = tk.StringVar()
        self.form_fields["excel"] = FormField(
            excel_row, "Excel 文件", self.vars["excel"],
            helper="支持 .xlsx 和 .xlsm；Linux 可用 LibreOffice 另存，或点“新建模板”直接生成",
        )
        self.form_fields["excel"].grid(row=0, column=0, sticky="ew")
        self.excel_button = PillButton(excel_row, "选择文件", self._choose_excel, variant="outline", width=96,
                                       bg=TOKENS.surface)
        self.excel_button.grid(row=0, column=1, sticky="s", padx=(TOKENS.space_2, 0), pady=(20, 0))
        self.excel_template_button = PillButton(excel_row, "新建模板", self._new_excel_template, variant="outline", width=96,
                                                bg=TOKENS.surface)
        self.excel_template_button.grid(row=0, column=2, sticky="s", padx=(TOKENS.space_2, 0), pady=(20, 0))

        date_row = tk.Frame(body, bg=TOKENS.surface)
        date_row.grid(row=4, column=0, sticky="ew", pady=(0, TOKENS.space_3))
        date_row.columnconfigure(0, weight=1)
        self.order_date = tk.StringVar()
        self.order_date_field = FormField(date_row, "目标日期", self.order_date,
                                          helper="留空默认今天；只允许选择今天或过去日期")
        self.order_date_field.grid(row=0, column=0, sticky="ew")
        self.order_date_field.entry.bind("<Button-1>", self._open_order_date_picker, add="+")
        self.order_date_button = tk.Button(date_row, text="日历", command=self._open_order_date_picker,
                                           relief="flat", bd=0, bg=TOKENS.surface_muted, fg=TOKENS.primary,
                                           activebackground=TOKENS.green_light, font=(FONT_FAMILY, 9, "bold"),
                                           padx=10, pady=7)
        self.order_date_button.grid(row=0, column=1, sticky="s", padx=(TOKENS.space_2, 0), pady=(20, 0))

        order_row = tk.Frame(body, bg=TOKENS.surface)
        order_row.grid(row=5, column=0, sticky="ew", pady=(0, TOKENS.space_3))
        order_row.columnconfigure(0, weight=1)
        tk.Label(order_row, text="待处理订单数", bg=TOKENS.surface, fg=TOKENS.text,
                 font=(FONT_FAMILY, 10, "bold"), anchor="w").grid(row=0, column=0, sticky="w")
        self.order_count = tk.StringVar(value="1")
        self.order_spinbox = tk.Spinbox(order_row, from_=1, to=9999, textvariable=self.order_count,
                                        width=7, relief="flat", bd=0, bg=TOKENS.surface_muted,
                                        fg=TOKENS.text, buttonbackground=TOKENS.ceramic,
                                        font=(FONT_FAMILY, 11), highlightthickness=1,
                                        highlightbackground=TOKENS.border, highlightcolor=TOKENS.focus)
        self.order_spinbox.grid(row=0, column=1, sticky="e", ipady=6)
        self.order_error = tk.Label(body, text="", bg=TOKENS.surface, fg=TOKENS.error,
                                    font=(FONT_FAMILY, 9), anchor="w")
        self.order_error.grid(row=6, column=0, sticky="ew")

        remember_row = tk.Frame(body, bg=TOKENS.surface)
        remember_row.grid(row=7, column=0, sticky="ew", pady=(TOKENS.space_2, TOKENS.space_3))
        self.remember = tk.BooleanVar(value=True)
        tk.Checkbutton(remember_row, text="保存到系统凭据管理器", variable=self.remember,
                       bg=TOKENS.surface, fg=TOKENS.text_soft, activebackground=TOKENS.surface,
                       selectcolor=TOKENS.green_light, font=(FONT_FAMILY, 10), anchor="w",
                       highlightthickness=0).pack(side="left")

        actions = tk.Frame(body, bg=TOKENS.surface)
        actions.grid(row=8, column=0, sticky="ew", pady=(TOKENS.space_2, TOKENS.space_3))
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        self.order_start_button = PillButton(actions, "开始处理", self.start, variant="primary")
        self.order_start_button.grid(row=0, column=0, sticky="ew", padx=(0, TOKENS.space_1))
        self.order_stop_button = PillButton(actions, "停止", self.stop, variant="danger")
        self.order_stop_button.grid(row=0, column=1, sticky="ew", padx=(TOKENS.space_1, 0))
        self.order_stop_button.set_state("disabled")

        tools = tk.Frame(body, bg=TOKENS.surface)
        self.tools = tools
        tools.grid(row=9, column=0, sticky="ew")
        self.tool_buttons = [
            PillButton(tools, "检查浏览器", self.install_browser, variant="outline"),
            PillButton(tools, "清除密码", self.clear_password, variant="outline"),
            PillButton(tools, "检查更新", lambda: self.check_for_updates(manual=True), variant="outline"),
        ]

    def _build_sss_form(self) -> None:
        body = self.sss_form
        body.columnconfigure(0, weight=1)

        fields = [("闪时送网址", "sss_url", "闪时送下单平台地址"),
                  ("闪时送账号", "sss_account", "用于登录闪时送平台"),
                  ("登录密码", "sss_password", "密码仅保存在系统凭据管理器中")]
        self.vars_sss: dict[str, tk.StringVar] = {key: tk.StringVar() for _, key, _ in fields if key != "sss_password"}
        self.sss_password = tk.StringVar()
        self.sss_form_fields: dict[str, FormField] = {}
        for row, (label, key, helper) in enumerate(fields, start=0):
            variable = self.sss_password if key == "sss_password" else self.vars_sss[key]
            field = FormField(body, label, variable, show="*" if key == "sss_password" else "", helper=helper)
            field.grid(row=row, column=0, sticky="ew", pady=(0, TOKENS.space_3))
            self.sss_form_fields[key] = field

        sss_excel_row = tk.Frame(body, bg=TOKENS.surface)
        self.sss_excel_row = sss_excel_row
        sss_excel_row.grid(row=3, column=0, sticky="ew", pady=(0, TOKENS.space_3))
        sss_excel_row.columnconfigure(0, weight=1)
        self.vars_sss["sss_excel"] = tk.StringVar()
        self.sss_form_fields["sss_excel"] = FormField(sss_excel_row, "订单 Excel 文件", self.vars_sss["sss_excel"], helper="午餐/晚餐两表，A=姓名 B=门牌号 C=电话；没有文件可点“新建模板”")
        self.sss_form_fields["sss_excel"].grid(row=0, column=0, sticky="ew")
        self.sss_excel_button = PillButton(sss_excel_row, "选择文件", self._choose_sss_excel, variant="outline", width=96,
                                           bg=TOKENS.surface)
        self.sss_excel_button.grid(row=0, column=1, sticky="s", padx=(TOKENS.space_2, 0), pady=(20, 0))
        self.sss_template_button = PillButton(sss_excel_row, "新建模板", self._new_sss_template, variant="outline", width=96,
                                              bg=TOKENS.surface)
        self.sss_template_button.grid(row=0, column=2, sticky="s", padx=(TOKENS.space_2, 0), pady=(20, 0))

        self.vars_sss["sss_product_name"] = tk.StringVar()
        self.sss_form_fields["sss_product_name"] = FormField(body, "商品名称", self.vars_sss["sss_product_name"], helper="下单时商品“名称”的默认值")
        self.sss_form_fields["sss_product_name"].grid(row=4, column=0, sticky="ew", pady=(0, TOKENS.space_3))
        self.vars_sss["sss_common_address"] = tk.StringVar()
        self.sss_form_fields["sss_common_address"] = FormField(body, "常用地址", self.vars_sss["sss_common_address"], helper="下单时选择的常用地址")
        self.sss_form_fields["sss_common_address"].grid(row=5, column=0, sticky="ew", pady=(0, TOKENS.space_3))

        remember_row = tk.Frame(body, bg=TOKENS.surface)
        remember_row.grid(row=6, column=0, sticky="ew", pady=(TOKENS.space_2, TOKENS.space_3))
        self.sss_remember = tk.BooleanVar(value=True)
        tk.Checkbutton(remember_row, text="保存到系统凭据管理器", variable=self.sss_remember,
                       bg=TOKENS.surface, fg=TOKENS.text_soft, activebackground=TOKENS.surface,
                       selectcolor=TOKENS.green_light, font=(FONT_FAMILY, 10), anchor="w",
                       highlightthickness=0).pack(side="left")

        actions = tk.Frame(body, bg=TOKENS.surface)
        actions.grid(row=7, column=0, sticky="ew", pady=(TOKENS.space_2, TOKENS.space_3))
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        self.sss_start_button = PillButton(actions, "开始下单", self.start, variant="primary")
        self.sss_start_button.grid(row=0, column=0, sticky="ew", padx=(0, TOKENS.space_1))
        self.sss_stop_button = PillButton(actions, "停止", self.stop, variant="danger")
        self.sss_stop_button.grid(row=0, column=1, sticky="ew", padx=(TOKENS.space_1, 0))
        self.sss_stop_button.set_state("disabled")

        tools = tk.Frame(body, bg=TOKENS.surface)
        tools.grid(row=8, column=0, sticky="ew")
        for column in range(2):
            tools.columnconfigure(column, weight=1)
        self.sss_tool_buttons = [
            PillButton(tools, "检查浏览器", self.install_browser, variant="outline"),
            PillButton(tools, "清除密码", self.clear_sss_password, variant="outline"),
        ]
        self.sss_tool_buttons[0].grid(row=0, column=0, sticky="ew", padx=(0, TOKENS.space_1))
        self.sss_tool_buttons[1].grid(row=0, column=1, sticky="ew", padx=(TOKENS.space_1, 0))

    def _on_config_body_configure(self, _event: tk.Event[tk.Misc]) -> None:
        if self._config_layout_job is None:
            self._config_layout_job = self.after(16, self._apply_config_layout_from_widget)

    def _apply_config_layout_from_widget(self) -> None:
        self._config_layout_job = None
        width = max(1, self.config_card.body.winfo_width())
        self._apply_config_layout(width)

    def _apply_config_layout(self, width: int) -> None:
        """Reflow the left form before its controls become too narrow."""
        compact = form_layout_mode(width) == "compact"
        if self._config_compact is not None and compact == self._config_compact:
            return
        self._config_compact = compact

        self.excel_row.columnconfigure(0, weight=1)
        self.excel_row.columnconfigure(1, weight=0)
        self.excel_row.columnconfigure(2, weight=0)
        self.form_fields["excel"].grid_configure(row=0, column=0, columnspan=1, sticky="ew")
        if compact:
            self.excel_button.grid_configure(row=1, column=0, columnspan=1, sticky="e",
                                             padx=0, pady=(TOKENS.space_2, 0))
            self.excel_template_button.grid_configure(row=1, column=1, columnspan=2, sticky="e",
                                                      padx=(TOKENS.space_2, 0), pady=(TOKENS.space_2, 0))
        else:
            self.excel_button.grid_configure(row=0, column=1, columnspan=1, sticky="s",
                                             padx=(TOKENS.space_2, 0), pady=(20, 0))
            self.excel_template_button.grid_configure(row=0, column=2, columnspan=1, sticky="s",
                                                      padx=(TOKENS.space_2, 0), pady=(20, 0))

        # 闪时送表单的“选择文件”按钮同样需要适配窄布局。
        if hasattr(self, "sss_excel_row"):
            self.sss_excel_row.columnconfigure(0, weight=1)
            self.sss_excel_row.columnconfigure(1, weight=0)
            self.sss_excel_row.columnconfigure(2, weight=0)
            self.sss_form_fields["sss_excel"].grid_configure(row=0, column=0, columnspan=1, sticky="ew")
            if compact:
                self.sss_excel_button.grid_configure(row=1, column=0, columnspan=1, sticky="e",
                                                     padx=0, pady=(TOKENS.space_2, 0))
                self.sss_template_button.grid_configure(row=1, column=1, columnspan=2, sticky="e",
                                                        padx=(TOKENS.space_2, 0), pady=(TOKENS.space_2, 0))
            else:
                self.sss_excel_button.grid_configure(row=0, column=1, columnspan=1, sticky="s",
                                                     padx=(TOKENS.space_2, 0), pady=(20, 0))
                self.sss_template_button.grid_configure(row=0, column=2, columnspan=1, sticky="s",
                                                        padx=(TOKENS.space_2, 0), pady=(20, 0))

        for column in range(3):
            self.tools.columnconfigure(column, weight=0 if compact else 1)
        if compact:
            self.tools.columnconfigure(0, weight=1)
            self.tools.columnconfigure(1, weight=1)
            self.tool_buttons[0].grid(row=0, column=0, columnspan=1, sticky="ew",
                                      padx=(0, TOKENS.space_1), pady=(0, TOKENS.space_1))
            self.tool_buttons[1].grid(row=0, column=1, columnspan=1, sticky="ew",
                                      padx=(TOKENS.space_1, 0), pady=(0, TOKENS.space_1))
            self.tool_buttons[2].grid(row=1, column=0, columnspan=2, sticky="ew",
                                      padx=0, pady=(TOKENS.space_1, 0))
        else:
            self.tool_buttons[0].grid(row=0, column=0, columnspan=1, sticky="ew",
                                      padx=(0, TOKENS.space_1), pady=0)
            self.tool_buttons[1].grid(row=0, column=1, columnspan=1, sticky="ew",
                                      padx=TOKENS.space_1, pady=0)
            self.tool_buttons[2].grid(row=0, column=2, columnspan=1, sticky="ew",
                                      padx=(TOKENS.space_1, 0), pady=0)

    def _build_log_card(self) -> None:
        body = self.log_card.body
        body.columnconfigure(0, weight=1)
        body.rowconfigure(4, weight=1)
        top = tk.Frame(body, bg=TOKENS.surface)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(0, weight=1)
        tk.Label(top, text="运行日志", bg=TOKENS.surface, fg=TOKENS.primary,
                 font=(FONT_FAMILY, 20, "bold"), anchor="w").grid(row=0, column=0, sticky="w")
        self.status_badge = StatusBadge(top, state="ready", bg=TOKENS.surface)
        self.status_badge.grid(row=0, column=1, sticky="e")
        tk.Label(body, text="实时记录订单处理、浏览器检查和更新状态。", bg=TOKENS.surface,
                 fg=TOKENS.text_soft, font=(FONT_FAMILY, 10), anchor="w").grid(row=1, column=0, sticky="w", pady=(TOKENS.space_1, TOKENS.space_3))

        search = tk.Frame(body, bg=TOKENS.surface)
        search.grid(row=2, column=0, sticky="ew", pady=(0, TOKENS.space_3))
        search.columnconfigure(0, weight=1)
        search_shell = tk.Frame(search, bg=TOKENS.border)
        search_shell.grid(row=0, column=0, sticky="ew", padx=(0, TOKENS.space_2))
        search_entry = tk.Entry(search_shell, textvariable=self._search_var, relief="flat", bd=0,
                                bg=TOKENS.surface_muted, fg=TOKENS.text, insertbackground=TOKENS.primary,
                                font=(FONT_FAMILY, 10), highlightthickness=0)
        search_entry.pack(fill="both", expand=True, padx=1, pady=1, ipady=7)
        search_entry.bind("<Return>", lambda _event: self.search_log())
        search.columnconfigure(0, weight=1)
        actions = tk.Frame(search, bg=TOKENS.surface)
        actions.grid(row=0, column=1, sticky="e")
        PillButton(actions, "搜索", self.search_log, variant="primary", width=76).pack(side="left", padx=(0, TOKENS.space_1))
        PillButton(actions, "清除", self.clear_log_search, variant="outline", width=76).pack(side="left")

        log_shell = tk.Frame(body, bg=TOKENS.house, bd=0, highlightthickness=0)
        log_shell.grid(row=4, column=0, sticky="nsew")
        log_shell.columnconfigure(0, weight=1)
        log_shell.rowconfigure(0, weight=1)
        self.log = tk.Text(log_shell, height=18, state="disabled", wrap="word", bg=TOKENS.house,
                           fg=TOKENS.text_on_dark, insertbackground=TOKENS.text_on_dark,
                           selectbackground=TOKENS.uplift, relief="flat", bd=0, padx=14, pady=14,
                           font=FONT_MONO, spacing1=2, spacing3=2)
        self.log.grid(row=0, column=0, sticky="nsew")
        scrollbar = tk.Scrollbar(log_shell, orient="vertical", command=self.log.yview,
                                 bg=TOKENS.uplift, troughcolor=TOKENS.house,
                                 activebackground=TOKENS.green_light, relief="flat", bd=0)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.log.configure(yscrollcommand=scrollbar.set)

    def _on_resize(self, event: tk.Event[tk.Misc] | None = None) -> None:
        if event is not None and getattr(event, "widget", self) is not self:
            return
        if self._resize_job is not None:
            try:
                self.after_cancel(self._resize_job)
            except tk.TclError:
                pass
        self._resize_job = self.after(50, self._reflow_window)

    def _reflow_window(self) -> None:
        self._resize_job = None
        mode = layout_mode(self.winfo_width())
        if mode == self._layout_mode:
            return
        self._layout_mode = mode
        self.split_pane.set_mode(mode)
        if mode == "split":
            self.split_pane.set_ratio(self._split_ratio)

    def _on_split_ratio_committed(self, ratio: float) -> None:
        self._split_ratio = clamp_split_ratio(ratio)
        self._saved_config.split_ratio = self._split_ratio
        try:
            self._saved_config.save()
        except OSError:
            # A read-only profile must not make dragging fail.
            pass

    def _set_status(self, state: str) -> None:
        self.status_badge.set(state)

    def _load_saved_values(self) -> None:
        config = self._saved_config
        self.vars["url"].set(config.target_url)
        self.vars["phone"].set(config.phone_number)
        self.vars["excel"].set(str(config.excel_path) if config.excel_path else "")
        self.order_date.set(config.order_date or _dt.date.today().isoformat())
        self.password.set(get_password(config.phone_number) or "")
        self.vars_sss["sss_url"].set(config.sss_url)
        self.vars_sss["sss_account"].set(config.sss_account)
        self.vars_sss["sss_excel"].set(str(config.sss_excel_path) if config.sss_excel_path else "")
        self.vars_sss["sss_product_name"].set(config.sss_product_name)
        self.vars_sss["sss_common_address"].set(config.sss_common_address)
        self.sss_password.set(get_sss_password(config.sss_account) or "")
        self._split_ratio = clamp_split_ratio(config.split_ratio)
        if config.excel_path and config.excel_path.is_file():
            self.form_fields["excel"].set_state("valid")
        if config.sss_excel_path and config.sss_excel_path.is_file():
            self.sss_form_fields["sss_excel"].set_state("valid")

    def _initial_dir(self, current: str) -> str | None:
        """文件对话框的初始目录：优先用当前已选文件的目录，否则用用户主目录。"""
        from pathlib import Path as _Path

        try:
            parent = _Path(current).expanduser().parent
            if current and parent.is_dir():
                return str(parent)
        except OSError:
            pass
        try:
            home = _Path.home()
            if home.is_dir():
                return str(home)
        except OSError:
            pass
        return None

    def _choose_excel(self) -> None:
        path = filedialog.askopenfilename(
            parent=self,
            title="选择 Excel 文件（.xlsx / .xlsm）",
            defaultextension=".xlsx",
            filetypes=list(EXCEL_FILETYPES),
            initialdir=self._initial_dir(self.vars["excel"].get().strip()),
        )
        if path:
            self.vars["excel"].set(path)
            selected = AppConfig(excel_path=path).excel_path
            if selected and selected.is_file() and selected.suffix.lower() in EXCEL_EXTS:
                self.form_fields["excel"].set_state("valid", "文件已选择")
            else:
                self.form_fields["excel"].set_state("invalid", "请选择 .xlsx 或 .xlsm 文件")

    def _choose_sss_excel(self) -> None:
        path = filedialog.askopenfilename(
            parent=self,
            title="选择闪时送订单 Excel（.xlsx / .xlsm）",
            defaultextension=".xlsx",
            filetypes=list(EXCEL_FILETYPES),
            initialdir=self._initial_dir(self.vars_sss["sss_excel"].get().strip()),
        )
        if path:
            self.vars_sss["sss_excel"].set(path)
            selected = AppConfig(sss_excel_path=path).sss_excel_path
            if selected and selected.is_file() and selected.suffix.lower() in EXCEL_EXTS:
                self.sss_form_fields["sss_excel"].set_state("valid", "文件已选择")
            else:
                self.sss_form_fields["sss_excel"].set_state("invalid", "请选择 .xlsx 或 .xlsm 文件")

    def _new_excel_template(self) -> None:
        """Linux 没有预装 Excel 时，一键生成排单用的 .xlsx 模板。"""
        from pathlib import Path as _Path

        target = filedialog.asksaveasfilename(
            parent=self,
            title="新建排单 Excel 模板",
            defaultextension=".xlsx",
            filetypes=list(EXCEL_FILETYPES),
            initialdir=self._initial_dir(self.vars["excel"].get().strip()),
            initialfile="排单.xlsx",
        )
        if not target:
            return
        dest = _Path(target)
        if dest.suffix.lower() not in EXCEL_EXTS:
            dest = dest.with_suffix(".xlsx")
        try:
            self._write_order_template(dest)
        except Exception as exc:  # 缺 openpyxl 等依赖时给明确提示
            messagebox.showerror("新建模板失败", f"无法写入模板文件：\n{exc}", parent=self)
            return
        self.vars["excel"].set(str(dest))
        self.form_fields["excel"].set_state("valid", "模板已生成")
        self._append(f"已生成排单模板：{dest}")

    def _new_sss_template(self) -> None:
        """生成闪时送下单用的《午餐/晚餐》两表模板。"""
        from pathlib import Path as _Path

        target = filedialog.asksaveasfilename(
            parent=self,
            title="新建闪时送订单模板",
            defaultextension=".xlsx",
            filetypes=list(EXCEL_FILETYPES),
            initialdir=self._initial_dir(self.vars_sss["sss_excel"].get().strip()),
            initialfile="闪时送.xlsx",
        )
        if not target:
            return
        dest = _Path(target)
        if dest.suffix.lower() not in EXCEL_EXTS:
            dest = dest.with_suffix(".xlsx")
        try:
            self._write_sss_template(dest)
        except Exception as exc:
            messagebox.showerror("新建模板失败", f"无法写入模板文件：\n{exc}", parent=self)
            return
        self.vars_sss["sss_excel"].set(str(dest))
        self.sss_form_fields["sss_excel"].set_state("valid", "模板已生成")
        self._append(f"已生成闪时送模板：{dest}")

    @staticmethod
    def _write_order_template(dest: object) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        # 订单处理是“按需建表”：缺哪张表程序会自动创建，这里预建常用表并写好表头，
        # Linux 用户无需安装 Excel 即可直接拿去用（LibreOffice/WPS 也可打开）。
        weekday_headers = ["单号", "姓名", "地址", "电话", "经济/豪华", "总餐次"]
        weekdays = ("周一", "周二", "周三", "周四", "周五", "周六", "周日")
        first = True
        for name in (*weekdays, "衣锦中餐", "衣锦晚餐", "医学院中餐", "医学院晚餐", "东湖中餐", "东湖晚餐"):
            ws = wb.active if first else wb.create_sheet(title=name)
            if first:
                ws.title = name
                first = False
            ws.append(weekday_headers)
            ws.append([])  # 第 2 行为占位行，数据从第 3 行开始（与程序读写规则一致）
        wb.save(str(dest))
        wb.close()

    @staticmethod
    def _write_sss_template(dest: object) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        first = True
        for name in ("午餐", "晚餐"):
            ws = wb.active if first else wb.create_sheet(title=name)
            if first:
                ws.title = name
                first = False
            ws["A1"] = "姓名"
            ws["B1"] = "门牌号"
            ws["C1"] = "电话"
            ws["D1"] = "送达时间"
            # 第 2 行占位，第 3 行起填数据；示例行方便 Linux 用户照格式填写。
            ws["A2"] = ""
            ws["A3"] = "张三"
            ws["B3"] = "1栋101"
            ws["C3"] = "13800000000"
            ws["D3"] = "11:00" if name == "午餐" else "17:00"
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 12
        wb.save(str(dest))
        wb.close()

    def _config(self) -> AppConfig:
        try:
            count = int(self.order_count.get())
            if count < 1:
                raise ValueError
        except ValueError as exc:
            raise ValueError("待处理订单数必须是大于等于 1 的整数") from exc
        return AppConfig(target_url=self.vars["url"].get().strip(), phone_number=self.vars["phone"].get().strip(),
                         excel_path=self.vars["excel"].get().strip(), order_date=self.order_date.get().strip(),
                         browser_mode="auto")

    def _config_sss(self) -> AppConfig:
        return AppConfig(
            sss_url=self.vars_sss["sss_url"].get().strip(),
            sss_account=self.vars_sss["sss_account"].get().strip(),
            sss_excel_path=self.vars_sss["sss_excel"].get().strip(),
            sss_product_name=self.vars_sss["sss_product_name"].get().strip(),
            sss_common_address=self.vars_sss["sss_common_address"].get().strip(),
            browser_mode="auto",
        )

    def _validate_form(self) -> tuple[AppConfig, int] | None:
        """Validate inputs and render semantic field states before starting."""
        for field in self.form_fields.values():
            field.set_state("neutral")
        self.order_error.configure(text="")
        try:
            config = self._config()
            count = int(self.order_count.get())
            if count < 1:
                raise ValueError
        except ValueError:
            self.order_error.configure(text="请输入大于等于 1 的整数", fg=TOKENS.error)
            self.order_spinbox.configure(highlightbackground=TOKENS.error, highlightcolor=TOKENS.error)
            return None

        valid = True
        if not config.url:
            self.form_fields["url"].set_state("invalid", "请输入管理网址")
            valid = False
        else:
            self.form_fields["url"].set_state("valid")
        if not config.phone:
            self.form_fields["phone"].set_state("invalid", "请输入手机号或账号")
            valid = False
        else:
            self.form_fields["phone"].set_state("valid")
        if not self.password.get():
            self.form_fields["password"].set_state("invalid", "请输入登录密码")
            valid = False
        else:
            self.form_fields["password"].set_state("valid")
        if not config.excel_path or not config.excel_path.is_file():
            self.form_fields["excel"].set_state("invalid", "请选择存在的 Excel 文件")
            valid = False
        elif config.excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            self.form_fields["excel"].set_state("invalid", "请选择 .xlsx 或 .xlsm 文件")
            valid = False
        else:
            self.form_fields["excel"].set_state("valid", "文件已准备")
        try:
            selected_date = parse_target_date(config.order_date)
        except ValueError as exc:
            self.order_date_field.set_state("invalid", str(exc))
            valid = False
        else:
            date_message = "未设置，默认今天" if not config.order_date else f"将筛选 {selected_date.isoformat()} 的订单"
            self.order_date_field.set_state("valid", date_message)
        if not valid:
            self._set_status("error")
            return None
        self.order_spinbox.configure(highlightbackground=TOKENS.border, highlightcolor=TOKENS.focus)
        return config, count

    def _open_order_date_picker(self, _event: tk.Event[tk.Misc] | None = None) -> str | None:
        try:
            selected = parse_target_date(self.order_date.get())
        except ValueError:
            selected = _dt.date.today()
        existing = getattr(self, "_order_date_popup", None)
        if existing is not None:
            try:
                existing.destroy()
            except tk.TclError:
                pass
        self._order_date_popup = _DatePickerPopup(self, self.order_date_field.entry, selected)
        return "break" if _event is not None else None

    def _set_order_date(self, value: _dt.date) -> None:
        self.order_date.set(value.isoformat())
        self.order_date_field.set_state("valid", f"将筛选 {value.isoformat()} 的订单")

    def _clear_order_date(self) -> None:
        self.order_date.set("")
        self.order_date_field.set_state("neutral", "留空默认今天；只允许选择今天或过去日期")

    def _validate_sss_form(self) -> AppConfig | None:
        """Validate 闪时送 inputs and render semantic field states before starting."""
        for field in self.sss_form_fields.values():
            field.set_state("neutral")
        config = self._config_sss()
        valid = True
        if not config.sss_url:
            self.sss_form_fields["sss_url"].set_state("invalid", "请输入闪时送网址")
            valid = False
        else:
            self.sss_form_fields["sss_url"].set_state("valid")
        if not config.sss_account:
            self.sss_form_fields["sss_account"].set_state("invalid", "请输入闪时送账号")
            valid = False
        else:
            self.sss_form_fields["sss_account"].set_state("valid")
        if not self.sss_password.get():
            self.sss_form_fields["sss_password"].set_state("invalid", "请输入登录密码")
            valid = False
        else:
            self.sss_form_fields["sss_password"].set_state("valid")
        if not config.sss_excel_path or not config.sss_excel_path.is_file():
            self.sss_form_fields["sss_excel"].set_state("invalid", "请选择存在的 Excel 文件")
            valid = False
        elif config.sss_excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            self.sss_form_fields["sss_excel"].set_state("invalid", "请选择 .xlsx 或 .xlsm 文件")
            valid = False
        else:
            self.sss_form_fields["sss_excel"].set_state("valid", "文件已准备")
        if not valid:
            self._set_status("error")
            return None
        return config

    def _append(self, text: str) -> None:
        self._log_lines.append(text.rstrip())
        self.log.configure(state="normal")
        self.log.insert("end", text.rstrip() + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def search_log(self) -> None:
        query = self._search_var.get().strip().lower()
        if not query:
            return
        # Repeated searches walk through all matches (wrapping around) instead
        # of stopping at the first one; a new query or 清除 restarts the scan.
        if getattr(self, "_search_query", None) != query:
            self._search_query = query
            self._search_index = -1
        total = len(self._log_lines)
        for offset in range(total):
            index = (self._search_index + 1 + offset) % total
            if query in self._log_lines[index].lower():
                self._search_index = index
                self.log.configure(state="normal")
                self.log.tag_remove("search_hit", "1.0", "end")
                start = f"{index + 1}.0"
                end = f"{index + 1}.end"
                self.log.tag_add("search_hit", start, end)
                self.log.tag_configure("search_hit", background="#fff2a8")
                self.log.see(start)
                self.log.configure(state="disabled")
                return
        self._search_index = -1
        messagebox.showinfo("搜索结果", f"未找到包含“{self._search_var.get().strip()}”的订单日志。")

    def clear_log_search(self) -> None:
        self._search_var.set("")
        self._search_query = None
        self._search_index = -1
        self.log.configure(state="normal")
        self.log.tag_remove("search_hit", "1.0", "end")
        self.log.configure(state="disabled")

    def _drain_events(self) -> None:
        try:
            while True:
                kind, value = self.events.get_nowait()
                if kind == "log":
                    self._append(value)
                elif kind == "done":
                    self._append(value)
                    self.start_button.set_state("normal")
                    self.stop_button.set_state("disabled")
                    self._set_status("success")
                    self.worker = None
                elif kind == "error":
                    self._append("错误: " + value)
                    messagebox.showerror("处理失败", value)
                    self.start_button.set_state("normal")
                    self.stop_button.set_state("disabled")
                    self._set_status("error")
                    self.worker = None
                elif kind == "browser_missing":
                    self._append(value)
                    self.start_button.set_state("normal")
                    self.stop_button.set_state("disabled")
                    self._set_status("error")
                    self.worker = None
                    self.choose_browser_download()
                elif kind == "update":
                    release = value
                    if isinstance(release, ReleaseInfo):
                        details = release.body or "（暂无更新说明）"
                        can_auto_install = ((os.name == "nt" or sys.platform.startswith("linux"))
                                            and getattr(sys, "frozen", False))
                        action = "是否立即下载并安装？" if can_auto_install else "是否打开 GitHub Release 下载页面？"
                        prompt = f"发现新版本 {release.tag_name}（当前版本 {__version__}）\n\n更新内容：\n{details}\n\n{action}"
                        if messagebox.askyesno("发现新版本", prompt):
                            if can_auto_install:
                                self._install_update(release)
                            elif release.html_url:
                                webbrowser.open(release.html_url)
                elif kind == "update_latest":
                    self._set_status("ready")
                    messagebox.showinfo("检查更新", f"当前已是最新版本（{__version__}）。")
                elif kind == "update_error":
                    self._append("检查更新失败：" + value)
                    self._set_status("error")
                elif kind == "update_progress":
                    downloaded, total = value
                    self._set_update_progress(downloaded, total)
                elif kind == "update_stage":
                    self._set_update_stage(str(value))
                elif kind == "update_install_error":
                    self._close_update_progress()
                    self._append("更新失败：" + value)
                    self._set_status("error")
                    if messagebox.askyesno("更新失败", f"{value}\n\n是否打开 GitHub Release 页面手动下载？"):
                        release = getattr(self, "_pending_release", None)
                        if isinstance(release, ReleaseInfo) and release.html_url:
                            webbrowser.open(release.html_url)
                elif kind == "update_installed":
                    self._append(value)
                    self._set_update_stage("正在重启")
                    self._set_status("success")
                    self._set_update_progress(1, 1)
                    messagebox.showinfo("更新完成", "更新已下载，点击确定后程序将关闭并自动重启。")
                    self._closing = True
                    self.destroy()
        except queue.Empty:
            pass
        self.after(100, self._drain_events)

    def _on_close(self) -> None:
        if not self.worker or not self.worker.is_alive():
            self.destroy()
            return
        if self._closing:
            return
        action = messagebox.askyesnocancel("正在处理", "任务仍在运行。点击“是”停止并关闭，点击“否”继续处理，点击“取消”返回。")
        if action is None or action is False:
            return
        self._closing = True
        self.stop_event.set()
        self._append("正在停止并清理浏览器，请稍候...")
        self.after(100, self._wait_for_worker_close)

    def _wait_for_worker_close(self) -> None:
        if self.worker and self.worker.is_alive():
            self.after(100, self._wait_for_worker_close)
        else:
            self.destroy()

    def start(self) -> None:
        mode = getattr(self, "_mode", "order")
        if mode == "sss":
            self._start_sss()
        else:
            self._start_order()

    def _start_order(self) -> None:
        validated = self._validate_form()
        if validated is None:
            return
        config, count = validated
        config.save()
        if self.remember.get():
            set_password(config.phone_number, self.password.get())
        self.stop_event.clear()
        self.start_button.set_state("disabled")
        self.stop_button.set_state("normal")
        self._set_status("running")
        self._append("开始处理订单...")
        self.worker = threading.Thread(target=self._run, args=(config, count, self.password.get()), daemon=True)
        self.worker.start()

    def _start_sss(self) -> None:
        config = self._validate_sss_form()
        if config is None:
            return
        config.save()
        if self.sss_remember.get():
            set_sss_password(config.sss_account, self.sss_password.get())
        self.stop_event.clear()
        self.start_button.set_state("disabled")
        self.stop_button.set_state("normal")
        self._set_status("running")
        self._append("开始闪时送下单...")
        self.worker = threading.Thread(target=self._run_sss, args=(config, self.sss_password.get()), daemon=True)
        self.worker.start()

    def _run(self, config: AppConfig, count: int, password: str) -> None:
        try:
            result = run_job(config, count, self.stop_event, lambda msg: self.events.put(("log", msg)), password=password,
                             order_decision_callback=self._order_decision,
                             save_decision_callback=self._save_decision)
            self.events.put(("done", f"处理完成：{result}"))
        except BrowserNotFoundError as exc:
            self.events.put(("browser_missing", str(exc)))
        except Exception as exc:
            self.events.put(("error", str(exc)))

    def _run_sss(self, config: AppConfig, password: str) -> None:
        try:
            result = run_sss_job(config, self.stop_event,
                                 lambda msg: self.events.put(("log", msg)),
                                 password=password,
                                 decision_callback=self._sss_decision)
            self.events.put(("done", f"闪时送下单完成：{result}"))
        except BrowserNotFoundError as exc:
            self.events.put(("browser_missing", str(exc)))
        except Exception as exc:
            self.events.put(("error", str(exc)))

    def stop(self) -> None:
        if not self.worker or not self.worker.is_alive():
            return
        action = messagebox.askyesnocancel("暂停处理", "是否停止当前任务？点击“是”停止，点击“否”继续处理，点击“取消”返回。")
        if action:
            self.stop_event.set()
            self._set_status("stopping")
            self._append("已请求停止，正在等待浏览器操作结束...")

    def _order_decision(self, code: str, error: str) -> str:
        result: queue.Queue[str] = queue.Queue(maxsize=1)
        def ask() -> None:
            choice = messagebox.askyesnocancel("订单定位失败", f"订单 {code} 定位失败：\n{error}\n\n是=重试，否=跳过，取消=停止")
            result.put("retry" if choice is True else "skip" if choice is False else "stop")
        self.after(0, ask)
        return result.get()

    def _sss_decision(self, identifier: str, error: str) -> str:
        result: queue.Queue[str] = queue.Queue(maxsize=1)
        def ask() -> None:
            choice = messagebox.askyesnocancel("下单失败", f"订单 {identifier} 创建失败：\n{error}\n\n是=重试，否=跳过，取消=停止")
            result.put("retry" if choice is True else "skip" if choice is False else "stop")
        self.after(0, ask)
        return result.get()

    def _save_decision(self, error: str) -> str:
        result: queue.Queue[str] = queue.Queue(maxsize=1)
        def ask() -> None:
            choice = messagebox.askretrycancel(
                "Excel 文件正在使用",
                "保存失败，Excel 文件可能正在被打开或占用。\n请关闭 Excel 文件后点击“重试保存”。\n\n" + error,
            )
            result.put("retry" if choice else "cancel")
        self.after(0, ask)
        return result.get()

    def install_browser(self) -> None:
        self._set_status("updating")
        self._append("正在检查浏览器...")
        threading.Thread(target=self._install_browser_worker, daemon=True).start()

    def _install_browser_worker(self) -> None:
        try:
            path = ensure_browser("auto")
            self.events.put(("log", f"浏览器可用：{path}"))
        except Exception as exc:
            self.events.put(("error", str(exc)))

    def choose_browser_download(self) -> None:
        choice = messagebox.askyesno("未安装浏览器", "打开 Microsoft Edge 下载页面吗？选择“否”将打开 Google Chrome 下载页面。")
        webbrowser.open("https://www.microsoft.com/edge/download" if choice else "https://www.google.com/chrome/")

    def check_for_updates(self, manual: bool = False) -> None:
        if getattr(self, "_update_checking", False):
            return
        self._update_checking = True
        self._set_status("updating")
        self._append("正在检查更新...")
        threading.Thread(target=self._check_updates_worker, args=(manual,), daemon=True).start()

    def _check_updates_worker(self, manual: bool) -> None:
        try:
            release = check_for_update()
            self.events.put(("update", release) if release else ("update_latest", "") if manual else ("log", "已是最新版本"))
        except UpdateError as exc:
            self.events.put(("update_error", str(exc)))
        finally:
            self._update_checking = False

    def _install_update(self, release: ReleaseInfo) -> None:
        self._pending_release = release
        self._append(f"获取更新清单完成，正在下载版本 {release.version}...")
        self._show_update_progress(release)
        threading.Thread(target=self._install_update_worker, args=(release,), daemon=True).start()

    def _install_update_worker(self, release: ReleaseInfo) -> None:
        try:
            download_and_install(
                release,
                progress_callback=lambda downloaded, total: self.events.put(("update_progress", (downloaded, total))),
                stage_callback=lambda stage: self.events.put(("update_stage", stage)),
            )
            self.events.put(("update_installed", "更新已下载，程序将重启"))
        except Exception as exc:
            self.events.put(("update_install_error", str(exc)))

    def _show_update_progress(self, release: ReleaseInfo) -> None:
        existing = getattr(self, "_update_progress_window", None)
        if existing and existing.winfo_exists():
            existing.lift()
            return
        dialog = tk.Toplevel(self)
        self._update_progress_window = dialog
        dialog.title("正在更新")
        dialog.geometry("520x210")
        dialog.resizable(False, False)
        dialog.configure(bg=TOKENS.canvas)
        dialog.transient(self)
        dialog.protocol("WM_DELETE_WINDOW", lambda: None)
        card = RoundedCard(dialog, parent_bg=TOKENS.canvas, padding=TOKENS.space_4)
        card.pack(fill="both", expand=True, padx=TOKENS.space_3, pady=TOKENS.space_3)
        content = card.body
        content.columnconfigure(0, weight=1)
        tk.Label(content, text=f"正在下载一口轻食 {release.tag_name}", bg=TOKENS.surface,
                 fg=TOKENS.primary, font=(FONT_FAMILY, 16, "bold"), anchor="w").pack(anchor="w")
        self._update_progress_text = tk.StringVar(value="获取更新清单…")
        tk.Label(content, textvariable=self._update_progress_text, bg=TOKENS.surface,
                 fg=TOKENS.text_soft, font=(FONT_FAMILY, 10), anchor="w").pack(anchor="w", pady=(TOKENS.space_2, TOKENS.space_2))
        self._update_progressbar = ttk.Progressbar(content, maximum=100, mode="indeterminate")
        self._update_progressbar.pack(fill="x")
        self._update_progressbar.start(12)
        tk.Label(content, text="下载完成后程序会自动关闭、替换并重新启动。", bg=TOKENS.surface,
                 fg=TOKENS.text_soft, font=(FONT_FAMILY, 9), anchor="w").pack(anchor="w", pady=(TOKENS.space_2, 0))
        dialog.grab_set()

    def _set_update_progress(self, downloaded: int, total: int | None) -> None:
        dialog = getattr(self, "_update_progress_window", None)
        if not dialog or not dialog.winfo_exists():
            return
        downloaded_mb = downloaded / (1024 * 1024)
        if downloaded == 0:
            self._update_progress_text.set("下载更新…")
            return
        if total:
            self._update_progressbar.stop()
            self._update_progressbar.configure(mode="determinate", value=min(downloaded * 100 / total, 100))
            percent = min(downloaded * 100 / total, 100)
            self._update_progress_text.set(f"下载更新… 已下载 {downloaded_mb:.1f} MB / {total / (1024 * 1024):.1f} MB（{percent:.0f}%）")
        else:
            self._update_progress_text.set(f"下载更新… 已下载 {downloaded_mb:.1f} MB")

    def _set_update_stage(self, stage: str) -> None:
        dialog = getattr(self, "_update_progress_window", None)
        if dialog and dialog.winfo_exists() and hasattr(self, "_update_progress_text"):
            self._update_progress_text.set(stage + "…")

    def _close_update_progress(self) -> None:
        dialog = getattr(self, "_update_progress_window", None)
        if dialog and dialog.winfo_exists():
            self._update_progressbar.stop()
            dialog.grab_release()
            dialog.destroy()

    def clear_password(self) -> None:
        phone = self.vars["phone"].get().strip()
        if phone:
            delete_password(phone)
        self.password.set("")
        self._append("已清除本机保存的密码")

    def clear_sss_password(self) -> None:
        account = self.vars_sss["sss_account"].get().strip()
        if account:
            delete_sss_password(account)
        self.sss_password.set("")
        self._append("已清除本机保存的闪时送密码")


def main() -> None:
    App().mainloop()
