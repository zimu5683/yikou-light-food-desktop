"""processing 纯函数的单元测试：收货人解析、楼栋判定与地址段提取。"""
from __future__ import annotations

import datetime as dt

from openpyxl import Workbook

from app.models import MealInfo, OrderInfo
from app.processing import (
    get_address_base_sheet_name,
    get_donghu_address_segment,
    get_first_empty_row,
    get_yijin_address_from_product_note,
    parse_meal_text,
    parse_receiver_info,
    write_order_row,
)


def test_parse_receiver_info_accepts_common_formats():
    assert parse_receiver_info("张三（13800000001）") == ("张三", "13800000001")
    assert parse_receiver_info("张三(13800000001)") == ("张三", "13800000001")
    assert parse_receiver_info("张三，13800000001") == ("张三", "13800000001")
    assert parse_receiver_info("张三:13800000001") == ("张三", "13800000001")
    assert parse_receiver_info("李四 13900000002") == ("李四", "13900000002")
    assert parse_receiver_info("王五") == ("王五", "")
    assert parse_receiver_info("") == ("", "")
    assert parse_receiver_info(None) == ("", "")


def test_address_base_sheet_maps_keywords_and_nonglin_road():
    assert get_address_base_sheet_name("联建1栋302") == "衣锦"
    assert get_address_base_sheet_name("lianjian 101") == "衣锦"
    assert get_address_base_sheet_name("衣锦校区") == "衣锦"
    assert get_address_base_sheet_name("医学院宿舍") == "医学院"
    assert get_address_base_sheet_name("东湖小区3栋") == "东湖"
    # 农林路默认归东湖，除非明确提到联建。
    assert get_address_base_sheet_name("农林路2号") == "东湖"
    assert get_address_base_sheet_name("农林路联建门口") == "衣锦"
    assert get_address_base_sheet_name("完全陌生的地址") is None


def test_donghu_segment_extracts_room_from_landmarks():
    assert get_donghu_address_segment("东湖大西12栋A101") == "A101"
    assert get_donghu_address_segment("小西3幢B202") == "B202"
    assert get_donghu_address_segment("东湖大西活动室") == "大西"
    assert get_donghu_address_segment("东湖小西") == "小西"
    assert get_donghu_address_segment("其他地址") == "其他地址"


def test_yijin_note_picks_cabinet_or_gate():
    assert get_yijin_address_from_product_note("备注：联建门口外卖柜自提") == "外卖柜"
    assert get_yijin_address_from_product_note("放校门口") == "校门口"
    assert get_yijin_address_from_product_note("") == "校门口"


def test_parse_meal_text_extracts_grade_and_count():
    meals = parse_meal_text("豪华轻食六餐x2（午餐）", "午餐")
    assert len(meals) == 1
    assert meals[0].total_meals == 6
    assert meals[0].grade == "豪华"
    assert meals[0].count == 2
    assert meals[0].meal_type == "午餐"

    single = parse_meal_text("经济轻食单点（晚餐）", "晚餐")
    assert single[0].total_meals == 1
    assert single[0].grade == "经济"

    plain = parse_meal_text("轻食套餐", "午餐")
    assert plain[0].meal_type == "午餐"
    assert plain[0].total_meals is None
    assert plain[0].count == 1


def test_write_order_row_appends_lunch_and_dinner_columns():
    wb = Workbook()
    ws = wb.active
    order = OrderInfo(order_no="W1", name="张三", address="大西A101", phone="13800000001")
    lunch = MealInfo(total_meals=6, grade="经济", count=1, meal_type="午餐")

    assert write_order_row(ws, order, lunch, "午餐") == 3
    assert [ws["A3"].value, ws["B3"].value, ws["C3"].value, ws["D3"].value,
            ws["E3"].value, ws["F3"].value] == ["W1", "张三", "大西A101", "13800000001", "经济", 6]

    dinner = MealInfo(total_meals=1, grade="豪华", count=1, meal_type="晚餐")
    assert write_order_row(ws, order, dinner, "晚餐") == 3
    assert [ws["G3"].value, ws["H3"].value, ws["K3"].value, ws["L3"].value] == \
        ["W1", "张三", "豪华", 1]

    assert write_order_row(ws, order, lunch, "午餐") == 4


def test_get_first_empty_row_skips_leading_placeholder_rows():
    ws = Workbook().active
    ws["A1"] = "表头"
    ws["A2"] = "占位"
    assert get_first_empty_row(ws) == 3


def test_historical_order_writes_to_a_dated_sheet():
    from app.automation import _write_order

    workbook = Workbook()
    order = OrderInfo(order_no="W2", name="张三", address="大西A101", phone="13800000001")
    meal = MealInfo(total_meals=6, grade="经济", count=1, meal_type="午餐")
    target_date = dt.date(2026, 9, 3)

    _write_order(workbook, order, meal, "午餐", target_date=target_date, today=dt.date(2026, 9, 4))
    sheet = workbook["2026年9月3日 周四"]

    assert [sheet.cell(1, column).value for column in range(1, 5)] == ["取单号", "姓名", "地址", "电话"]
    assert [sheet.cell(2, column).value for column in range(1, 5)] == ["W2", "张三", "大西A101", "13800000001"]
    assert sheet.cell(2, 8).value == 1  # 周四
    assert sheet.cell(2, 12).value == "午餐"


def test_group_orders_by_pick_keeps_api_order_and_skips_empty_pick_no():
    from app.automation import _group_orders_by_pick

    rows = [
        {"order_id": "3", "pick_no": "W1", "date": dt.date(2026, 9, 3)},
        {"order_id": "2", "pick_no": "", "date": dt.date(2026, 9, 3)},
        {"order_id": "1", "pick_no": "W8", "date": dt.date(2026, 7, 6)},
    ]
    grouped = _group_orders_by_pick(rows)
    assert set(grouped) == {"W1", "W8"}
    assert grouped["W1"][0]["order_id"] == "3"


def test_group_orders_by_pick_groups_repeated_pick_numbers_across_days():
    """取单号跨天重复：同号不同日期的行都保留，顺序与新单在前一致。"""
    from app.automation import _group_orders_by_pick

    rows = [
        {"order_id": "30", "pick_no": "W2", "date": dt.date(2026, 9, 3)},
        {"order_id": "7", "pick_no": "W2", "date": dt.date(2026, 7, 6)},
        {"order_id": "6", "pick_no": "W1", "date": dt.date(2026, 7, 6)},
    ]
    grouped = _group_orders_by_pick(rows)
    assert [r["order_id"] for r in grouped["W2"]] == ["30", "7"]
    assert [r["order_id"] for r in grouped["W1"]] == ["6"]


def test_filter_rows_by_date_keeps_only_target_date():
    from app.automation import _filter_rows_by_date

    target = dt.date(2026, 9, 7)
    rows = [
        {"order_id": "1", "pick_no": "W8", "date": target},
        {"order_id": "2", "pick_no": "W7", "date": dt.date(2026, 9, 6)},
        {"order_id": "3", "pick_no": "W3", "date": target},
        {"order_id": "4", "pick_no": "W1", "date": None},
    ]
    filtered = _filter_rows_by_date(rows, target)
    assert [r["order_id"] for r in filtered] == ["1", "3"]


def test_order_numbers_for_date_returns_descending_existing_numbers():
    from app.automation import _order_numbers_for_date

    rows_by_pick = {
        "W8": [{"order_id": "1", "date": dt.date(2026, 9, 7)}],
        "W3": [{"order_id": "2", "date": dt.date(2026, 9, 7)}],
        "W1": [{"order_id": "3", "date": dt.date(2026, 9, 7)}],
    }
    # 留空/0 表示全部
    assert _order_numbers_for_date(rows_by_pick, None) == [8, 3, 1]
    assert _order_numbers_for_date(rows_by_pick, 0) == [8, 3, 1]
    # 指定数量时只保留当天存在且不超过该编号的
    assert _order_numbers_for_date(rows_by_pick, 3) == [3, 1]
    # 没有订单时为空
    assert _order_numbers_for_date({}, None) == []
